from __future__ import annotations

import argparse
import csv
import hashlib
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Final, cast

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT: Final = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE: Final = PROJECT_ROOT / "sources" / "Datos_Ema_Serrana_INN.xlsx"
DEFAULT_OUTPUT_DIR: Final = PROJECT_ROOT / "data"
SHEET_FILENAMES: Final = {
    "Diseño": "diseno.csv",
    "Ensayo": "ensayo.csv",
    "Manejo": "manejo.csv",
    "Datos_MS": "datos_ms.csv",
    "Datos_Rto": "datos_rto.csv",
    "Calidad": "calidad.csv",
    "Estimacion_M1R4_16set": "estimacion_m1r4_16set.csv",
}


@dataclass(frozen=True)
class Arguments:
    source: Path
    output_dir: Path
    check: bool


@dataclass(frozen=True)
class SheetSummary:
    name: str
    filename: str
    rows: int
    columns: int
    formula_cells: int
    formulas_without_cached_values: int


def parse_args() -> Arguments:
    parser = argparse.ArgumentParser(
        description="Export every worksheet to Git-friendly UTF-8 CSV files."
    )
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--check",
        action="store_true",
        help="Verify that existing CSV files match the workbook without rewriting them.",
    )
    namespace = parser.parse_args()
    return Arguments(
        source=cast(Path, namespace.source),
        output_dir=cast(Path, namespace.output_dir),
        check=cast(bool, namespace.check),
    )


def serialize_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == time.min:
            return value.date().isoformat()
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_or_check_csv(
    path: Path, rows: Sequence[Sequence[str]], *, check: bool
) -> None:
    expected = [list(row) for row in rows]
    if not check:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle, lineterminator="\n")
            writer.writerows(expected)

    if not path.is_file():
        raise FileNotFoundError(f"Expected CSV does not exist: {path}")

    with path.open(encoding="utf-8", newline="") as handle:
        actual = list(csv.reader(handle))
    if actual != expected:
        raise RuntimeError(f"CSV differs from workbook export: {path}")


def collect_sheet_rows(
    formula_sheet: Worksheet, value_sheet: Worksheet, filename: str
) -> tuple[list[list[str]], list[list[str]], SheetSummary]:
    csv_rows: list[list[str]] = []
    formula_rows: list[list[str]] = []
    formula_cells = 0
    formulas_without_cached_values = 0

    for row_number in range(1, formula_sheet.max_row + 1):
        csv_row: list[str] = []
        for column_number in range(1, formula_sheet.max_column + 1):
            formula_cell = formula_sheet.cell(row_number, column_number)
            value_cell = value_sheet.cell(row_number, column_number)
            csv_row.append(serialize_value(value_cell.value))

            if formula_cell.data_type == "f":
                formula_cells += 1
                if value_cell.value is None:
                    formulas_without_cached_values += 1
                formula_rows.append(
                    [
                        formula_sheet.title,
                        formula_cell.coordinate,
                        serialize_value(formula_cell.value),
                        serialize_value(value_cell.value),
                    ]
                )
        csv_rows.append(csv_row)

    summary = SheetSummary(
        name=formula_sheet.title,
        filename=filename,
        rows=formula_sheet.max_row,
        columns=formula_sheet.max_column,
        formula_cells=formula_cells,
        formulas_without_cached_values=formulas_without_cached_values,
    )
    return csv_rows, formula_rows, summary


def source_label(source: Path) -> str:
    try:
        return source.relative_to(PROJECT_ROOT).as_posix()
    except ValueError:
        return source.as_posix()


def build_manifest_rows(
    source: Path, summaries: Sequence[SheetSummary]
) -> list[list[str]]:
    source_digest = sha256_file(source)
    rows = [
        [
            "source_file",
            "source_sha256",
            "sheet",
            "csv_file",
            "rows",
            "columns",
            "formula_cells",
            "formulas_without_cached_values",
        ]
    ]
    for summary in summaries:
        rows.append(
            [
                source_label(source),
                source_digest,
                summary.name,
                summary.filename,
                str(summary.rows),
                str(summary.columns),
                str(summary.formula_cells),
                str(summary.formulas_without_cached_values),
            ]
        )
    return rows


def export_workbook(source: Path, output_dir: Path, *, check: bool) -> None:
    if not source.is_file():
        raise FileNotFoundError(f"Workbook does not exist: {source}")

    formula_book = load_workbook(source, data_only=False, read_only=False)
    value_book = load_workbook(source, data_only=True, read_only=False)
    try:
        if formula_book.sheetnames != value_book.sheetnames:
            raise RuntimeError(
                "Formula and value workbook views have different sheets."
            )

        unknown_sheets = set(formula_book.sheetnames) - set(SHEET_FILENAMES)
        if unknown_sheets:
            names = ", ".join(sorted(unknown_sheets))
            raise RuntimeError(f"Add CSV filename mappings for new sheets: {names}")

        formula_rows: list[list[str]] = [["sheet", "cell", "formula", "cached_value"]]
        summaries: list[SheetSummary] = []

        for sheet_name in formula_book.sheetnames:
            formula_sheet = formula_book[sheet_name]
            value_sheet = value_book[sheet_name]
            filename = SHEET_FILENAMES[sheet_name]
            csv_rows, sheet_formula_rows, summary = collect_sheet_rows(
                formula_sheet, value_sheet, filename
            )
            write_or_check_csv(output_dir / filename, csv_rows, check=check)
            formula_rows.extend(sheet_formula_rows)
            summaries.append(summary)

        write_or_check_csv(output_dir / "formulas.csv", formula_rows, check=check)

        manifest_rows = build_manifest_rows(source, summaries)
        write_or_check_csv(output_dir / "manifest.csv", manifest_rows, check=check)
    finally:
        formula_book.close()
        value_book.close()


def main() -> None:
    arguments = parse_args()
    export_workbook(
        arguments.source.resolve(),
        arguments.output_dir.resolve(),
        check=arguments.check,
    )


if __name__ == "__main__":
    main()
