"""Contract tests for the canonical CSV/JSON to XLSX reconstruction."""

# pyright: reportUnknownArgumentType=false, reportUnknownMemberType=false
# pyright: reportUnknownVariableType=false

from __future__ import annotations

import csv
import json
import re
import shutil
import tempfile
import unittest
from pathlib import Path
from typing import Any, cast
from unittest.mock import patch

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.workbook import Workbook

import festuca_analysis.xlsx_rebuild as rebuild_module
from festuca_analysis.xlsx_rebuild import (
    DATASET_LAYOUT,
    DOCUMENT_TABLES,
    FIRST_DATA_ROW,
    SHEET_ORDER,
    rebuild_workbook,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"

RECORDED_SHEETS = (
    "Diseño",
    "Metadatos",
    "Cronograma",
    "Suelo",
    "Manejo",
    "Agua",
    "MS registrada",
    "Cosecha registrada",
    "Calidad registrada",
)

EXPECTED_ROWS = {
    "experimental_design.csv": 24,
    "experiment_metadata.csv": 24,
    "field_timeline.csv": 18,
    "soil_analysis.csv": 8,
    "field_management.csv": 11,
    "water_inputs.csv": 11,
    "dry_matter_recorded.csv": 154,
    "dry_matter_calculated.csv": 154,
    "harvest_recorded.csv": 48,
    "harvest_calculated.csv": 48,
    "quality_recorded.csv": 152,
    "quality_calculated.csv": 152,
    "missing_quality_estimate.csv": 1,
}


def _header(spec: dict[str, Any]) -> str:
    unit = spec.get("unit")
    name = str(spec["display_name_es"])
    return f"{name} ({unit})" if unit else name


def _headers(workbook: Workbook, sheet_name: str) -> list[str]:
    worksheet = workbook[sheet_name]
    return [str(cell.value) for cell in worksheet[4]]


def _table_row_count(workbook: Workbook, sheet_name: str) -> int:
    table = next(iter(workbook[sheet_name].tables.values()))
    _, min_row, _, max_row = range_boundaries(table.ref)
    return cast(int, max_row) - cast(int, min_row)


class WorkbookRebuildTests(unittest.TestCase):
    temporary_directory: tempfile.TemporaryDirectory[str]
    output: Path
    workbook: Workbook
    manifest: dict[str, Any]
    formulas: dict[str, Any]

    @classmethod
    def setUpClass(cls) -> None:
        cls.temporary_directory = tempfile.TemporaryDirectory()
        cls.output = Path(cls.temporary_directory.name) / "canonical.xlsx"
        rebuild_workbook(DATA_DIR, cls.output)
        cls.workbook = load_workbook(cls.output, data_only=False)
        cls.manifest = cast(
            dict[str, Any],
            json.loads((DATA_DIR / "manifest.json").read_text(encoding="utf-8")),
        )
        cls.formulas = cast(
            dict[str, Any],
            json.loads((DATA_DIR / "formulas.json").read_text(encoding="utf-8")),
        )

    @classmethod
    def tearDownClass(cls) -> None:
        cls.workbook.close()
        cls.temporary_directory.cleanup()

    def test_sheet_order_tables_and_exact_row_counts(self) -> None:
        self.assertEqual(self.workbook.sheetnames, list(SHEET_ORDER))
        observed_table_names: list[str] = []
        for worksheet in self.workbook.worksheets:
            tables = list(worksheet.tables.values())
            self.assertEqual(len(tables), 1, worksheet.title)
            observed_table_names.append(tables[0].displayName)
            self.assertFalse(tables[0].totalsRowShown)
            self.assertEqual(worksheet.freeze_panes, "B5")
            self.assertFalse(worksheet.sheet_view.showGridLines)
            self.assertIsNone(worksheet.auto_filter.ref)
        self.assertEqual(len(observed_table_names), len(set(observed_table_names)))
        self.assertEqual(
            set(observed_table_names),
            set(DOCUMENT_TABLES.values())
            | {table_name for _, table_name in DATASET_LAYOUT.values()},
        )
        for filename, expected in EXPECTED_ROWS.items():
            sheet_name, _ = DATASET_LAYOUT[filename]
            self.assertEqual(_table_row_count(self.workbook, sheet_name), expected)
        self.assertEqual(_table_row_count(self.workbook, "Índice"), 13)
        self.assertEqual(_table_row_count(self.workbook, "Fórmulas"), 29)

    def test_spanish_headers_and_technical_dictionary(self) -> None:
        datasets = cast(dict[str, Any], self.manifest["datasets"])
        dictionary = self.workbook["Diccionario"]
        dictionary_rows = {
            (str(row[1].value), str(row[2].value), str(row[3].value))
            for row in dictionary.iter_rows(min_row=FIRST_DATA_ROW)
        }
        for filename, dataset in datasets.items():
            sheet_name, _ = DATASET_LAYOUT[filename]
            visible_headers = set(_headers(self.workbook, sheet_name))
            columns = cast(dict[str, dict[str, Any]], dataset["columns"])
            for technical_name, spec in columns.items():
                self.assertIn(_header(spec), visible_headers)
                self.assertIn(
                    (filename, technical_name, str(spec["display_name_es"])),
                    dictionary_rows,
                )

    def test_recorded_sheets_and_formula_catalog_are_formula_free(self) -> None:
        for sheet_name in RECORDED_SHEETS:
            worksheet = self.workbook[sheet_name]
            self.assertFalse(
                any(
                    cell.data_type == "f"
                    for row in worksheet.iter_rows(min_row=FIRST_DATA_ROW)
                    for cell in row
                ),
                sheet_name,
            )
        formula_sheet = self.workbook["Fórmulas"]
        template_column = (
            _headers(self.workbook, "Fórmulas").index("Plantilla Excel") + 1
        )
        for row in range(FIRST_DATA_ROW, formula_sheet.max_row + 1):
            cell = formula_sheet.cell(row=row, column=template_column)
            self.assertEqual(cell.data_type, "s")
            self.assertTrue(str(cell.value).startswith("="))

    def test_all_declared_targets_are_live_formulas(self) -> None:
        datasets = cast(dict[str, Any], self.manifest["datasets"])
        targets = {
            (str(formula["output_file"]), str(formula["output_column"]))
            for formula in cast(list[dict[str, Any]], self.formulas["formulas"])
        }
        self.assertEqual(len(targets), 29)
        for filename, technical_name in targets:
            sheet_name, _ = DATASET_LAYOUT[filename]
            worksheet = self.workbook[sheet_name]
            spec = cast(dict[str, Any], datasets[filename]["columns"][technical_name])
            column_index = _headers(self.workbook, sheet_name).index(_header(spec)) + 1
            for row_index in range(FIRST_DATA_ROW, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                self.assertEqual(
                    cell.data_type,
                    "f",
                    f"{sheet_name}!{cell.coordinate} ({technical_name})",
                )
                self.assertNotIn("#REF!", str(cell.value))
                self.assertNotIn("XLOOKUP", str(cell.value))

        for sheet_name in (
            "MS calculada",
            "Cosecha calculada",
            "Calidad calculada",
            "Calidad faltante",
        ):
            table = next(iter(self.workbook[sheet_name].tables.values()))
            worksheet = self.workbook[sheet_name]
            for column_index, _ in enumerate(table.tableColumns, start=1):
                first_cell = worksheet.cell(FIRST_DATA_ROW, column_index)
                if first_cell.data_type == "f":
                    formula_text = str(first_cell.value)
                    if "#This Row" in formula_text:
                        self.assertIn(f"{table.displayName}[[#This Row]", formula_text)

    def test_calculated_views_link_context_without_cycles(self) -> None:
        expected_sources = {
            "MS calculada": "tblMSRegistrada",
            "Cosecha calculada": "tblCosechaRegistrada",
            "Calidad calculada": "tblCalidadRegistrada",
        }
        for sheet_name, source_table in expected_sources.items():
            worksheet = self.workbook[sheet_name]
            formulas = [
                str(cell.value)
                for cell in worksheet[FIRST_DATA_ROW]
                if cell.data_type == "f"
            ]
            self.assertTrue(any(source_table in formula for formula in formulas))
        harvest_formulas = [
            str(cell.value)
            for cell in self.workbook["Cosecha calculada"][FIRST_DATA_ROW]
            if cell.data_type == "f"
        ]
        quality_formulas = [
            str(cell.value)
            for cell in self.workbook["Calidad calculada"][FIRST_DATA_ROW]
            if cell.data_type == "f"
        ]
        self.assertTrue(any("tblMSCalculada" in value for value in harvest_formulas))
        self.assertTrue(
            any("tblCalidadFaltante" in value for value in quality_formulas)
        )
        self.assertTrue(
            any(
                "INDEX(tblCalidadFaltante[Nitrógeno estimado (%)],MATCH(" in value
                for value in quality_formulas
            )
        )
        self.assertFalse(
            any(
                re.search(r"tblCalidadCalculada\[(?!\[)", value)
                for value in quality_formulas
            )
        )

    def test_native_types_formats_and_explicit_blanks(self) -> None:
        design = self.workbook["Diseño"]
        self.assertEqual(design["B5"].data_type, "n")
        self.assertIsInstance(design["B5"].value, int)

        timeline = self.workbook["Cronograma"]
        date_column = _headers(self.workbook, "Cronograma").index("Fecha") + 1
        self.assertEqual(timeline.cell(FIRST_DATA_ROW, date_column).data_type, "d")
        self.assertEqual(
            timeline.cell(FIRST_DATA_ROW, date_column).number_format, "yyyy-mm-dd"
        )
        boolean_column = (
            _headers(self.workbook, "Cronograma").index("Orden intradía conocido") + 1
        )
        boolean_cells = [
            timeline.cell(row, boolean_column)
            for row in range(FIRST_DATA_ROW, timeline.max_row + 1)
            if timeline.cell(row, boolean_column).value is not None
        ]
        self.assertTrue(boolean_cells)
        self.assertTrue(all(cell.data_type == "b" for cell in boolean_cells))

        metadata = self.workbook["Metadatos"]
        metadata_headers = _headers(self.workbook, "Metadatos")
        parameter_column = metadata_headers.index("Parámetro") + 1
        value_column = metadata_headers.index("Valor") + 1
        year_row = next(
            row
            for row in range(FIRST_DATA_ROW, metadata.max_row + 1)
            if metadata.cell(row, parameter_column).value == "experiment_year"
        )
        self.assertEqual(metadata.cell(year_row, value_column).value, 2025)
        self.assertIsInstance(metadata.cell(year_row, value_column).value, int)

        dry = self.workbook["MS registrada"]
        tiller_column = (
            _headers(self.workbook, "MS registrada").index(
                "Macollos en 30 cm (macollos/30 cm)"
            )
            + 1
        )
        self.assertTrue(
            any(
                dry.cell(row, tiller_column).value is None
                for row in range(FIRST_DATA_ROW, dry.max_row + 1)
            )
        )

    def test_water_totals_and_missing_quality_results(self) -> None:
        water = self.workbook["Agua"]
        headers = _headers(self.workbook, "Agua")
        irrigation_column = headers.index("Riego suplementario (mm)") + 1
        rainfall_column = headers.index("Precipitación (mm)") + 1
        self.assertEqual(
            sum(
                float(cast(Any, water.cell(row, irrigation_column).value))
                for row in range(FIRST_DATA_ROW, water.max_row + 1)
            ),
            165.0,
        )
        self.assertEqual(
            sum(
                float(cast(Any, water.cell(row, rainfall_column).value))
                for row in range(FIRST_DATA_ROW, water.max_row + 1)
            ),
            1176.0,
        )

        with (DATA_DIR / "missing_quality_estimate.csv").open(
            encoding="utf-8", newline=""
        ) as handle:
            estimate = cast(dict[str, str], next(iter(csv.DictReader(handle))))
        treatment_count = float(estimate["treatment_count"])
        block_count = float(estimate["block_count"])

        def dbca(prefix: str) -> float:
            return (
                treatment_count * float(estimate[f"{prefix}_block_total"])
                + block_count * float(estimate[f"{prefix}_treatment_total"])
                - float(estimate[f"{prefix}_grand_total"])
            ) / ((treatment_count - 1) * (block_count - 1))

        self.assertAlmostEqual(dbca("n"), 2.8688484862162937, places=12)
        self.assertAlmostEqual(dbca("adf"), 40.77623836505554, places=12)
        self.assertAlmostEqual(dbca("ndf"), 69.57708690222144, places=12)
        biomass = float(estimate["biomass_kg_ha"])
        n_value = dbca("n")
        n_accumulated = biomass * n_value / 100.0
        critical_n = 4.8 * (biomass / 1000.0) ** -0.32
        self.assertAlmostEqual(n_accumulated, 201.8808679750406, places=10)
        self.assertAlmostEqual(n_value / critical_n, 1.1159131393270323, places=12)

        missing_sheet = self.workbook["Calidad faltante"]
        missing_headers = _headers(self.workbook, "Calidad faltante")
        n_formula = missing_sheet.cell(
            FIRST_DATA_ROW, missing_headers.index("Nitrógeno estimado (%)") + 1
        )
        self.assertIn(
            "tblCalidadFaltante[[#This Row],[Número de tratamientos]]"
            "*tblCalidadFaltante[[#This Row],[Total de N del bloque (%)]]",
            str(n_formula.value),
        )
        self.assertNotIn(
            "tblCalidadFaltante[[#This Row],[Número de bloques]]"
            "*tblCalidadFaltante[[#This Row],[Total de N del bloque (%)]]",
            str(n_formula.value),
        )

    def test_calculation_mode_and_conditional_highlights(self) -> None:
        self.assertEqual(self.workbook.calculation.calcMode, "auto")
        self.assertTrue(self.workbook.calculation.fullCalcOnLoad)
        self.assertTrue(self.workbook.calculation.forceFullCalc)
        for sheet_name in (
            "MS calculada",
            "Calidad registrada",
            "Calidad calculada",
            "Calidad faltante",
        ):
            self.assertGreater(len(self.workbook[sheet_name].conditional_formatting), 0)

    def test_overwrite_force_lock_and_semantic_check(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "derived.xlsx"
            rebuild_workbook(DATA_DIR, output)
            rebuild_workbook(DATA_DIR, output, check=True)
            with self.assertRaises(FileExistsError):
                rebuild_workbook(DATA_DIR, output)
            rebuild_workbook(DATA_DIR, output, force=True)
            lock = output.with_name(f"~${output.name}")
            lock.touch()
            with self.assertRaisesRegex(RuntimeError, "Close Microsoft Excel"):
                rebuild_workbook(DATA_DIR, output, force=True)

    def test_semantic_drift_is_detected(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "derived.xlsx"
            rebuild_workbook(DATA_DIR, output)
            workbook = load_workbook(output)
            workbook["Índice"]["A1"] = "Alterado"
            workbook.save(output)
            workbook.close()
            with self.assertRaisesRegex(ValueError, "no coincide semánticamente"):
                rebuild_workbook(DATA_DIR, output, check=True)

    def test_force_replacement_is_atomic_on_build_failure(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "derived.xlsx"
            rebuild_workbook(DATA_DIR, output)
            original = output.read_bytes()
            with (
                patch.object(
                    rebuild_module, "_build_workbook", side_effect=RuntimeError("boom")
                ),
                self.assertRaisesRegex(RuntimeError, "boom"),
            ):
                rebuild_workbook(DATA_DIR, output, force=True)
            self.assertEqual(output.read_bytes(), original)

    def test_unknown_formula_fails_closed_without_output(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            copied_data = Path(directory) / "data"
            shutil.copytree(DATA_DIR, copied_data)
            formulas_path = copied_data / "formulas.json"
            payload = cast(
                dict[str, Any],
                json.loads(formulas_path.read_text(encoding="utf-8")),
            )
            cast(list[dict[str, Any]], payload["formulas"])[0]["id"] = "unknown"
            formulas_path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            output = Path(directory) / "should_not_exist.xlsx"
            with self.assertRaisesRegex(ValueError, "Fórmula desconocida"):
                rebuild_workbook(copied_data, output)
            self.assertFalse(output.exists())


if __name__ == "__main__":
    unittest.main()
