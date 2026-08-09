"""Workbook-first data loading and provenance for the Festuca experiment.

The XLSX workbook is the only source of observed values used by the analysis.
This module distinguishes recorded measurements, workbook formulas, estimates,
and quantities derived by the analysis.  It deliberately does not import result
values from the thesis document or from previously generated CSV files.
"""

from __future__ import annotations

# Pandas and openpyxl expose partially typed workbook/dataframe boundaries.
# pyright: reportUnknownArgumentType=false, reportUnknownMemberType=false
# pyright: reportUnknownVariableType=false
import hashlib
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Final, Literal, cast

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PROJECT_ROOT: Final = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK: Final = PROJECT_ROOT / "sources" / "Datos_Ema_Serrana_INN.xlsx"
TREATMENT_PATTERN: Final = re.compile(r"^M[0-5]$")

DataStatus = Literal[
    "recorded",
    "recorded_method_not_encoded",
    "calculated_in_workbook",
    "reported_derived_without_formula",
    "estimated_in_workbook",
    "missing",
    "analysis_derived",
    "metadata",
]
DryMatterPolicy = Literal["recorded", "ratio", "exclude"]


@dataclass(frozen=True)
class ExperimentSpec:
    """Design and management information parsed from the workbook."""

    workbook_path: Path
    source_sha256: str
    experiment_year: int
    treatments: tuple[str, ...]
    sectors: tuple[str, ...]
    blocks: tuple[str, ...]
    repetitions: int
    plot_area_m2: float
    row_spacing_m: float
    biomass_sample_area_m2: float
    harvest_sample_area_m2: float
    experimental_n_total_kg_ha: float
    applications_per_treatment: int
    dose_per_application_kg_ha: float
    schedule: pd.DataFrame
    management: pd.DataFrame
    water_monthly: pd.DataFrame
    water_period_totals: pd.DataFrame
    source_audit: pd.DataFrame


@dataclass(frozen=True)
class ExperimentData:
    """Analysis-ready tables plus their source and lineage metadata."""

    spec: ExperimentSpec
    longitudinal: pd.DataFrame
    harvest: pd.DataFrame
    seed_weight_long: pd.DataFrame
    baseline_biomass: pd.DataFrame
    baseline_tillers: pd.DataFrame
    variable_lineage: pd.DataFrame
    qa: pd.DataFrame


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def locate_workbook(
    workbook_path: Path | str | None = None,
    *,
    project_root: Path | None = None,
) -> Path:
    """Resolve the workbook without consulting generated CSV exports."""

    if workbook_path is not None:
        explicit = Path(workbook_path).expanduser().resolve()
        if not explicit.is_file():
            raise FileNotFoundError(f"No existe el libro indicado: {explicit}")
        return explicit

    root = (project_root or PROJECT_ROOT).resolve()
    candidates = [
        root / "sources" / DEFAULT_WORKBOOK.name,
        root / DEFAULT_WORKBOOK.name,
        Path("/mnt/data") / DEFAULT_WORKBOOK.name,
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate.resolve()
    attempted = ", ".join(str(candidate) for candidate in candidates)
    raise FileNotFoundError(
        f"No se encontró {DEFAULT_WORKBOOK.name!r}. Se probó: {attempted}"
    )


def _strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(
        character for character in normalized if not unicodedata.combining(character)
    )


def _normalize_text(value: object) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    return " ".join(str(value).strip().split())


def _normalize_treatment(value: object) -> str:
    text = _normalize_text(value).upper().replace("MO", "M0")
    return text


def _normalize_sector(value: object) -> str:
    text = _strip_accents(_normalize_text(value)).casefold()
    mapping = {"secano": "Secano", "riego": "Riego"}
    return mapping.get(text, _normalize_text(value))


def _normalize_block(value: object) -> str:
    return _normalize_text(value).upper()


def _coerce_numeric(frame: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    result = frame.copy()
    for column in columns:
        if column in result.columns:
            result[column] = pd.to_numeric(result[column], errors="coerce")
    return result


def _parse_partial_date(value: object, *, year: int) -> pd.Timestamp | None:
    if value is None:
        return None
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return pd.Timestamp(value)
    if isinstance(value, (int, float)) and not pd.isna(value):
        # Excel serial dates use 1899-12-30 as the origin in pandas/openpyxl.
        return pd.Timestamp("1899-12-30") + pd.to_timedelta(float(value), unit="D")

    text = _strip_accents(_normalize_text(value)).casefold()
    if not text or text in {"_", "-", "sin fecha"}:
        return None

    month_lookup = {
        "ene": 1,
        "enero": 1,
        "feb": 2,
        "febrero": 2,
        "mar": 3,
        "marzo": 3,
        "abr": 4,
        "abril": 4,
        "may": 5,
        "mayo": 5,
        "jun": 6,
        "junio": 6,
        "jul": 7,
        "julio": 7,
        "ago": 8,
        "agosto": 8,
        "sep": 9,
        "set": 9,
        "septiembre": 9,
        "setiembre": 9,
        "oct": 10,
        "octubre": 10,
        "nov": 11,
        "noviembre": 11,
        "dic": 12,
        "diciembre": 12,
    }
    match = re.fullmatch(r"(\d{1,2})\s*[-/]?\s*([a-z]+)", text)
    if match:
        day = int(match.group(1))
        month = month_lookup.get(match.group(2))
        if month is None:
            raise ValueError(f"Mes no reconocido en fecha parcial: {value!r}")
        return pd.Timestamp(year=year, month=month, day=day)

    parsed = pd.to_datetime(str(value), errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        raise ValueError(f"No se pudo interpretar la fecha del libro: {value!r}")
    parsed = pd.Timestamp(parsed)
    if parsed.year == 1900:
        parsed = parsed.replace(year=year)
    return parsed


def _month_number(value: object) -> int:
    text = _strip_accents(_normalize_text(value)).casefold()
    names = {
        "enero": 1,
        "febrero": 2,
        "marzo": 3,
        "abril": 4,
        "mayo": 5,
        "junio": 6,
        "julio": 7,
        "agosto": 8,
        "setiembre": 9,
        "septiembre": 9,
        "octubre": 10,
        "noviembre": 11,
        "diciembre": 12,
    }
    if text not in names:
        raise ValueError(f"Mes no reconocido: {value!r}")
    return names[text]


def _formula_coordinates(workbook_path: Path) -> dict[str, set[str]]:
    book = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        result: dict[str, set[str]] = {}
        for sheet_name in book.sheetnames:
            coordinates: set[str] = set()
            for row in book[sheet_name].iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        coordinates.add(cell.coordinate)
            result[sheet_name] = coordinates
        return result
    finally:
        book.close()


def _cell_status(
    *,
    formula_coordinates: dict[str, set[str]],
    sheet: str,
    column_number: int,
    excel_row: int,
    semantically_derived: bool,
) -> DataStatus:
    coordinate = f"{get_column_letter(column_number)}{excel_row}"
    if coordinate in formula_coordinates.get(sheet, set()):
        return "calculated_in_workbook"
    if semantically_derived:
        return "reported_derived_without_formula"
    return "recorded"


def _parse_schedule(value_sheet: object, *, experiment_year: int) -> pd.DataFrame:
    sheet = value_sheet["Ensayo"]  # type: ignore[index]
    rows: list[dict[str, object]] = []
    for row in range(2, 8):
        treatment = _normalize_treatment(sheet.cell(row, 6).value)
        if not TREATMENT_PATTERN.fullmatch(treatment):
            continue
        first = _parse_partial_date(sheet.cell(row, 7).value, year=experiment_year)
        second = _parse_partial_date(sheet.cell(row, 8).value, year=experiment_year)
        rows.append(
            {
                "treatment": treatment,
                "first_application": first,
                "second_application": second,
                "source_sheet": "Ensayo",
                "source_range": f"F{row}:H{row}",
            }
        )
    schedule = pd.DataFrame(rows).sort_values("treatment").reset_index(drop=True)
    expected = {f"M{i}" for i in range(6)}
    observed = set(schedule["treatment"])
    if observed != expected:
        raise ValueError(
            "La tabla estructurada de tratamientos en Ensayo no contiene M0–M5: "
            f"{sorted(observed)}"
        )
    return schedule


def _parse_management_and_water(
    value_sheet: object,
    *,
    experiment_year: int,
    study_start: pd.Timestamp,
    study_end: pd.Timestamp,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    sheet = value_sheet["Manejo"]  # type: ignore[index]
    management_rows: list[dict[str, object]] = []
    water_rows: list[dict[str, object]] = []
    for row in range(14, 25):
        management_month = sheet.cell(row, 1).value
        management_text = sheet.cell(row, 2).value
        if management_month is not None:
            month_number = _month_number(management_month)
            management_rows.append(
                {
                    "year": experiment_year,
                    "month": month_number,
                    "month_label": _normalize_text(management_month),
                    "management": _normalize_text(management_text) or pd.NA,
                    "source_sheet": "Manejo",
                    "source_row": row,
                }
            )

        water_month = sheet.cell(row, 6).value
        if water_month is not None:
            month_number = _month_number(water_month)
            irrigation = pd.to_numeric(sheet.cell(row, 7).value, errors="coerce")
            rainfall = pd.to_numeric(sheet.cell(row, 8).value, errors="coerce")
            water_rows.append(
                {
                    "year": experiment_year,
                    "month": month_number,
                    "month_label": _normalize_text(water_month),
                    "rainfall_mm": float(rainfall) if pd.notna(rainfall) else 0.0,
                    "supplemental_irrigation_mm": (
                        float(irrigation) if pd.notna(irrigation) else 0.0
                    ),
                    "source_sheet": "Manejo",
                    "source_row": row,
                }
            )

    management = pd.DataFrame(management_rows)
    water = pd.DataFrame(water_rows)
    water["month_start"] = pd.to_datetime(
        {
            "year": water["year"],
            "month": water["month"],
            "day": np.ones(len(water), dtype=int),
        }
    )
    water["included_in_study_months"] = water["month_start"].between(
        study_start.to_period("M").to_timestamp(),
        study_end.to_period("M").to_timestamp(),
        inclusive="both",
    )
    water["irrigated_sector_input_mm"] = (
        water["rainfall_mm"] + water["supplemental_irrigation_mm"]
    )

    period = water.loc[water["included_in_study_months"]]
    totals = pd.DataFrame(
        [
            {
                "sector": "Secano",
                "rainfall_mm": float(period["rainfall_mm"].sum()),
                "supplemental_irrigation_mm": 0.0,
                "gross_input_mm": float(period["rainfall_mm"].sum()),
                "aggregation": "meses calendario que intersectan el ensayo",
                "study_start": study_start,
                "study_end": study_end,
            },
            {
                "sector": "Riego",
                "rainfall_mm": float(period["rainfall_mm"].sum()),
                "supplemental_irrigation_mm": float(
                    period["supplemental_irrigation_mm"].sum()
                ),
                "gross_input_mm": float(period["irrigated_sector_input_mm"].sum()),
                "aggregation": "meses calendario que intersectan el ensayo",
                "study_start": study_start,
                "study_end": study_end,
            },
        ]
    )
    return management, water, totals


def _parse_harvest_geometry(value_sheet: object, *, row_spacing_m: float) -> float:
    """Derive harvested ground area from the workbook description.

    ``Datos_Rto!D3`` can state both total linear metres and an explanatory
    parenthesis such as ``(1 metro, 2 surcos)``.  The parenthetical values are
    preferred because multiplying the leading total by the row count would
    double-count the sampled length.
    """

    sheet = value_sheet["Datos_Rto"]  # type: ignore[index]
    raw_description = sheet["D3"].value
    description = _strip_accents(_normalize_text(raw_description)).casefold()

    parenthetical = re.search(
        r"\(\s*(\d+(?:[.,]\d+)?)\s*metros?\s*[,;x]\s*" r"(\d+)\s*surcos?\s*\)",
        description,
    )
    if parenthetical:
        length_per_row_m = float(parenthetical.group(1).replace(",", "."))
        row_count = int(parenthetical.group(2))
        return length_per_row_m * row_count * row_spacing_m

    total_linear_metres = re.search(
        r"(?:corte|muestra)\s+de\s+(\d+(?:[.,]\d+)?)\s*metros?",
        description,
    )
    if total_linear_metres:
        return float(total_linear_metres.group(1).replace(",", ".")) * row_spacing_m

    length_match = re.search(r"(\d+(?:[.,]\d+)?)\s*metros?", description)
    rows_match = re.search(r"(\d+)\s*surcos?", description)
    if length_match and rows_match:
        length_per_row_m = float(length_match.group(1).replace(",", "."))
        row_count = int(rows_match.group(1))
        return length_per_row_m * row_count * row_spacing_m

    raise ValueError(
        "No se pudo derivar el área de cosecha desde Datos_Rto!D3: "
        f"{raw_description!r}"
    )


def _agenda_audit_rows(
    value_sheet: object, experiment_year: int
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    ensayo = value_sheet["Ensayo"]  # type: ignore[index]
    for row in range(19, 34):
        agenda_date = ensayo.cell(row, 1).value
        agenda_text = _normalize_text(ensayo.cell(row, 2).value)
        if not agenda_text:
            continue
        parsed = _parse_partial_date(agenda_date, year=experiment_year)
        if pd.notna(parsed) and pd.Timestamp(parsed).year != experiment_year:
            rows.append(
                {
                    "severity": "warning",
                    "issue": "agenda_year_conflict",
                    "location": f"Ensayo!A{row}:B{row}",
                    "detail": (
                        "La fecha de la agenda cae fuera del año declarado del ensayo; "
                        "la tabla estructurada F:H se usa como calendario canónico."
                    ),
                    "observed": pd.Timestamp(parsed),
                }
            )

    return rows


def _application_sampling_audit_rows(
    schedule: pd.DataFrame, sample_dates: tuple[pd.Timestamp, ...]
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    sample_set = set(sample_dates)
    for record in schedule.itertuples(index=False):
        for application_number, application_date in enumerate(
            [record.first_application, record.second_application], start=1
        ):
            if (
                pd.notna(application_date)
                and pd.Timestamp(cast(Any, application_date)) in sample_set
            ):
                rows.append(
                    {
                        "severity": "warning",
                        "issue": "application_and_sampling_same_date",
                        "location": str(record.source_range),
                        "detail": (
                            "Aplicación y muestreo comparten fecha; el orden intradía no "
                            "está codificado en el libro. No se interpreta el muestreo como "
                            "respuesta posterior a esa aplicación sin bitácora de campo."
                        ),
                        "observed": (
                            f"{record.treatment}, aplicación {application_number}, "
                            f"{pd.Timestamp(cast(Any, application_date)).date()}"
                        ),
                    }
                )

    return rows


def _source_audit(
    *,
    value_sheet: object,
    schedule: pd.DataFrame,
    sample_dates: tuple[pd.Timestamp, ...],
    experiment_year: int,
    formula_coordinates: dict[str, set[str]],
) -> pd.DataFrame:
    rows = _agenda_audit_rows(value_sheet, experiment_year)
    rows.extend(_application_sampling_audit_rows(schedule, sample_dates))

    manejo = value_sheet["Manejo"]  # type: ignore[index]
    april_text = _normalize_text(manejo["B17"].value)
    if april_text:
        rows.append(
            {
                "severity": "warning",
                "issue": "common_n_active_dose_unresolved",
                "location": "Manejo!B17",
                "detail": (
                    "El libro registra masa de producto, pero no codifica su fracción de N. "
                    "No se transforma a kg N ha-1 ni se suma al N experimental."
                ),
                "observed": april_text,
            }
        )

    calidad = value_sheet["Calidad"]  # type: ignore[index]
    estimated_rows: list[int] = []
    for row in range(2, calidad.max_row + 1):
        origin = _strip_accents(_normalize_text(calidad.cell(row, 14).value)).casefold()
        if "estim" in origin:
            estimated_rows.append(row)
    if estimated_rows:
        rows.append(
            {
                "severity": "warning",
                "issue": "estimated_quality_values_present",
                "location": "Calidad!N:N",
                "detail": (
                    "Las filas marcadas como estimadas se conservan en columnas de auditoría, "
                    "pero se excluyen del análisis primario de mediciones."
                ),
                "observed": ", ".join(str(row) for row in estimated_rows),
            }
        )

    ms_formula_count = sum(
        coordinate.startswith("L")
        for coordinate in formula_coordinates.get("Datos_MS", set())
    )
    rows.append(
        {
            "severity": "info",
            "issue": "biomass_storage_is_mixed",
            "location": "Datos_MS!L:L",
            "detail": (
                "KgMS/ha aparece como fórmula en algunas filas y como valor literal en otras. "
                "El análisis lo recalcula uniformemente desde peso verde, %MS y distancia "
                "entre hileras."
            ),
            "observed": ms_formula_count,
        }
    )

    return pd.DataFrame(rows)


def _variable_lineage() -> pd.DataFrame:
    rows = [
        (
            "Fecha, condición, tratamiento, bloque, muestra",
            "metadata",
            "workbook",
            "Identificadores registrados.",
        ),
        ("Peso verde por metro", "recorded", "Datos_MS", "Entrada para biomasa."),
        (
            "Peso verde de submuestra",
            "recorded",
            "Datos_MS",
            "Entrada para auditoría de %MS.",
        ),
        ("Peso seco", "recorded", "Datos_MS", "Entrada para auditoría de %MS."),
        (
            "%MS",
            "recorded_method_not_encoded",
            "Datos_MS",
            "El valor está registrado; el procedimiento exacto no está codificado.",
        ),
        (
            "KgMS/ha registrado",
            "calculated_in_workbook / reported_derived_without_formula",
            "Datos_MS",
            "La forma de almacenamiento varía por fila.",
        ),
        (
            "Biomasa usada",
            "analysis_derived",
            "Peso verde, %MS y geometría",
            "Se recalcula con una sola identidad para todas las filas.",
        ),
        (
            "%N, %FDA, %FDN",
            "recorded / estimated_in_workbook",
            "Calidad",
            "El campo Origen del dato y el tipo de celda separan medición de estimación.",
        ),
        (
            "KgN/ha del libro",
            "calculated_in_workbook / estimated_in_workbook",
            "Calidad",
            "Fórmula del libro conservada solo para auditoría.",
        ),
        (
            "INN del libro",
            "calculated_in_workbook / estimated_in_workbook",
            "Calidad",
            "Fórmula histórica del libro conservada solo para auditoría.",
        ),
        (
            "N acumulado usado",
            "analysis_derived",
            "Biomasa recalculada y %N primario",
            "Producto determinista; la fila estimada no entra al análisis primario.",
        ),
        (
            "INN usado",
            "analysis_derived",
            "Biomasa recalculada, %N primario y curva crítica",
            "Índice determinista; se mantiene una curva de sensibilidad.",
        ),
        (
            "Panojas, peso sucio, peso limpio",
            "recorded",
            "Datos_Rto",
            "Mediciones primitivas de cosecha.",
        ),
        ("Tres pesos de 100 semillas", "recorded", "Datos_Rto", "Réplicas técnicas."),
        (
            "Peso de mil semillas",
            "calculated_in_workbook / analysis_derived",
            "Datos_Rto",
            "Promedio de réplicas × 10; se recalcula y audita.",
        ),
        (
            "Densidad de panojas",
            "analysis_derived",
            "Panojas y área cosechada",
            "Transformación geométrica.",
        ),
        (
            "Rendimientos limpio y sucio",
            "analysis_derived",
            "Masas y área cosechada",
            "Transformación geométrica.",
        ),
        (
            "Semillas estimadas por panoja",
            "analysis_derived",
            "Peso limpio, PMS y panojas",
            "Reconstrucción, no conteo independiente.",
        ),
        (
            "Índice de cosecha",
            "analysis_derived",
            "Rendimiento limpio y biomasa final",
            "Cociente determinista.",
        ),
        (
            "EAN",
            "analysis_derived",
            "Rendimiento y M0 del mismo bloque",
            "Transformación del rendimiento; uso descriptivo.",
        ),
        (
            "Productividad aparente del agua",
            "analysis_derived",
            "Rendimiento y entradas brutas mensuales",
            "No equivale a agua consumida por el cultivo.",
        ),
    ]
    return pd.DataFrame(
        rows, columns=["variable", "status", "source", "interpretation"]
    )


def _build_qa(
    *,
    spec: ExperimentSpec,
    longitudinal: pd.DataFrame,
    harvest: pd.DataFrame,
    sample_dates: tuple[pd.Timestamp, ...],
) -> pd.DataFrame:
    expected_plots = len(spec.treatments) * spec.repetitions * len(spec.sectors)
    expected_longitudinal = expected_plots * len(sample_dates)
    rows: list[dict[str, object]] = []

    def add(
        check: str,
        observed: object,
        expected: object,
        passed: bool,
        severity: str = "error",
    ) -> None:
        rows.append(
            {
                "check": check,
                "observed": observed,
                "expected": expected,
                "passes": bool(passed),
                "severity": severity,
            }
        )

    add(
        "treatments in schedule",
        len(spec.schedule),
        len(spec.treatments),
        len(spec.schedule) == len(spec.treatments),
    )
    add("harvest rows", len(harvest), expected_plots, len(harvest) == expected_plots)
    add(
        "longitudinal rows",
        len(longitudinal),
        expected_longitudinal,
        len(longitudinal) == expected_longitudinal,
    )
    add(
        "unique harvest plots",
        harvest["plot_id"].nunique(),
        expected_plots,
        harvest["plot_id"].nunique() == expected_plots,
    )
    add(
        "duplicate harvest plots",
        int(harvest["plot_id"].duplicated().sum()),
        0,
        not harvest["plot_id"].duplicated().any(),
    )
    add(
        "duplicate plot-date rows",
        int(longitudinal.duplicated(["plot_id", "date"]).sum()),
        0,
        not longitudinal.duplicated(["plot_id", "date"]).any(),
    )
    date_counts = longitudinal.groupby("plot_id", observed=True)["date"].nunique()
    add(
        "dates per plot (minimum)",
        int(date_counts.min()) if not date_counts.empty else 0,
        len(sample_dates),
        bool(not date_counts.empty and date_counts.min() == len(sample_dates)),
    )
    max_w1000_difference = float(
        (harvest["w1000_g"] - harvest["w1000_workbook_g"]).abs().max()
    )
    add(
        "max |PMS recomputed - workbook|",
        max_w1000_difference,
        "<= 1e-10 g",
        bool(max_w1000_difference <= 1e-10),
    )
    add(
        "missing measured N values",
        int(longitudinal["n_pct"].isna().sum()),
        "reported, not imputed in primary analysis",
        True,
        severity="info",
    )
    add(
        "dry-matter records flagged by dynamic rule",
        int(longitudinal["dm_issue"].sum()),
        "reported for verification",
        True,
        severity="info",
    )
    measured_q = longitudinal.loc[
        longitudinal["quality_status"].eq("recorded"),
        "q_workbook_recompute_difference",
    ].dropna()
    add(
        "max |N accumulated recomputed - workbook| for measured rows",
        float(measured_q.abs().max()) if not measured_q.empty else np.nan,
        "reported as reconciliation diagnostic",
        True,
        severity="info",
    )
    measured_nni = longitudinal.loc[
        longitudinal["quality_status"].eq("recorded"),
        "nni_sensitivity_workbook_difference",
    ].dropna()
    add(
        "max |sensitivity NNI recomputed - workbook| for measured rows",
        float(measured_nni.abs().max()) if not measured_nni.empty else np.nan,
        "reported as reconciliation diagnostic",
        True,
        severity="info",
    )
    return pd.DataFrame(rows)


def _apply_dry_matter_policy(frame: pd.DataFrame, policy: DryMatterPolicy) -> pd.Series:
    used = frame["%MS"].copy()
    if policy == "ratio":
        used.loc[frame["dm_issue"]] = frame.loc[frame["dm_issue"], "dm_ratio_pct"]
    elif policy == "exclude":
        used.loc[frame["dm_issue"]] = np.nan
    return used


def load_experiment_data(
    workbook_path: Path | str | None = None,
    *,
    project_root: Path | None = None,
    dry_matter_policy: DryMatterPolicy = "recorded",
    include_estimated_quality: bool = False,
    nni_primary_coefficient: float = 3.93,
    nni_primary_exponent: float = -0.42,
    nni_sensitivity_coefficient: float = 4.8,
    nni_sensitivity_exponent: float = -0.32,
    dm_absolute_difference_threshold: float = 5.0,
    dm_relative_difference_threshold: float = 0.20,
) -> ExperimentData:
    """Load and reconstruct the experiment from the XLSX workbook.

    Parameters controlling NNI are scientific assumptions rather than observed
    data.  They are explicit so a notebook can display them as methodology.
    """

    if dry_matter_policy not in {"recorded", "ratio", "exclude"}:
        raise ValueError("dry_matter_policy debe ser recorded, ratio o exclude")

    path = locate_workbook(workbook_path, project_root=project_root)
    formula_coordinates = _formula_coordinates(path)
    value_book = load_workbook(path, data_only=True, read_only=False)
    try:
        ensayo = value_book["Ensayo"]
        experiment_year = int(ensayo["B6"].value)
        repetitions = int(ensayo["B13"].value)
        plot_area_m2 = float(ensayo["B10"].value)
        experimental_n_total = float(ensayo["B8"].value)
        application_count = int(ensayo["B9"].value)
        dose_per_application = experimental_n_total / application_count
        schedule = _parse_schedule(value_book, experiment_year=experiment_year)

        ms_sheet = value_book["Datos_MS"]
        row_spacing_m = float(ms_sheet["K3"].value) / 100.0
        biomass_sample_area_m2 = row_spacing_m
        harvest_sample_area_m2 = _parse_harvest_geometry(
            value_book,
            row_spacing_m=row_spacing_m,
        )

        raw_ms = pd.read_excel(path, sheet_name="Datos_MS", header=4)
        raw_quality = pd.read_excel(path, sheet_name="Calidad", header=0)
        raw_harvest = pd.read_excel(path, sheet_name="Datos_Rto", header=4)

        ms = raw_ms.copy()
        ms["_excel_row"] = np.arange(6, 6 + len(ms))
        ms["Condición"] = ms["Condición"].map(_normalize_sector)
        ms["Tratamiento"] = ms["Tratamiento"].map(_normalize_treatment)
        ms["Repetición"] = ms["Repetición"].map(_normalize_block)
        ms["Fecha"] = pd.to_datetime(ms["Fecha"], errors="coerce")
        ms["Muestra"] = pd.to_numeric(ms["Muestra"], errors="coerce").astype("Int64")
        ms = _coerce_numeric(
            ms,
            [
                "Macollos/30 cm",
                "Peso verde (1m)",
                "Peso verde (muestra)",
                "Banjeda",
                "Peso Seco",
                "%MS",
                "KgMS/ha",
                "Macollos/m2",
            ],
        )
        ms["dm_ratio_pct"] = 100.0 * ms["Peso Seco"] / ms["Peso verde (muestra)"]
        ms["dm_abs_difference_pp"] = (ms["%MS"] - ms["dm_ratio_pct"]).abs()
        ms["dm_relative_difference"] = ms["dm_abs_difference_pp"] / ms[
            "%MS"
        ].abs().replace(0.0, np.nan)
        ms["dm_issue"] = (
            ms["dm_abs_difference_pp"].ge(dm_absolute_difference_threshold)
            & ms["dm_relative_difference"].ge(dm_relative_difference_threshold)
        ).fillna(False)
        ms["dm_pct_used"] = _apply_dry_matter_policy(ms, dry_matter_policy)

        ms["biomass_kg_ha_recomputed"] = (
            ms["Peso verde (1m)"]
            * (ms["dm_pct_used"] / 100.0)
            * 10.0
            / biomass_sample_area_m2
        )
        ms["biomass_recompute_difference"] = (
            ms["biomass_kg_ha_recomputed"] - ms["KgMS/ha"]
        )
        ms["kgms_workbook_status"] = [
            _cell_status(
                formula_coordinates=formula_coordinates,
                sheet="Datos_MS",
                column_number=12,
                excel_row=int(row),
                semantically_derived=True,
            )
            for row in ms["_excel_row"]
        ]
        ms["dm_pct_status"] = "recorded_method_not_encoded"

        baseline_ms = ms.loc[
            ~ms["Tratamiento"].str.fullmatch(TREATMENT_PATTERN, na=False)
        ].copy()
        experimental_ms = ms.loc[
            ms["Tratamiento"].str.fullmatch(TREATMENT_PATTERN, na=False)
        ].copy()

        quality = raw_quality.copy()
        quality["_excel_row"] = np.arange(2, 2 + len(quality))
        quality["Condición"] = quality["Condición"].map(_normalize_sector)
        quality["Tratamiento"] = quality["Tratamiento"].map(_normalize_treatment)
        quality["Repetición"] = quality["Repetición"].map(_normalize_block)
        quality["Fecha"] = pd.to_datetime(quality["Fecha"], errors="coerce")
        quality["Muestra"] = pd.to_numeric(quality["Muestra"], errors="coerce").astype(
            "Int64"
        )
        quality = _coerce_numeric(
            quality, ["%MS", "KgMS/ha", "% N", "% FDA ", "% FDN ", "KgN/ha", "INN"]
        )
        quality["origin_normalized"] = quality["Origen del dato"].map(
            lambda value: _strip_accents(_normalize_text(value)).casefold()
        )
        quality["quality_status"] = np.select(
            [
                quality["origin_normalized"].str.contains("medid", na=False),
                quality["origin_normalized"].str.contains("estim", na=False),
            ],
            ["recorded", "estimated_in_workbook"],
            default="missing",
        )
        quality["n_pct_recorded"] = quality["% N"]
        quality["n_pct_primary"] = quality["% N"]
        if not include_estimated_quality:
            quality.loc[quality["quality_status"].ne("recorded"), "n_pct_primary"] = (
                np.nan
            )

        quality_columns = {
            "n_pct_cell_status": (9, False),
            "adf_cell_status": (10, False),
            "ndf_cell_status": (11, False),
            "q_workbook_cell_status": (12, True),
            "nni_workbook_cell_status": (13, True),
        }
        for status_column, (
            column_number,
            semantically_derived,
        ) in quality_columns.items():
            statuses: list[DataStatus] = []
            for excel_row, origin_status in zip(
                quality["_excel_row"],
                quality["quality_status"],
                strict=True,
            ):
                if origin_status == "estimated_in_workbook":
                    statuses.append("estimated_in_workbook")
                else:
                    statuses.append(
                        _cell_status(
                            formula_coordinates=formula_coordinates,
                            sheet="Calidad",
                            column_number=column_number,
                            excel_row=int(excel_row),
                            semantically_derived=semantically_derived,
                        )
                    )
            quality[status_column] = statuses

        key = ["Fecha", "Muestra", "Condición", "Tratamiento", "Repetición"]
        longitudinal = experimental_ms[
            key
            + [
                "_excel_row",
                "%MS",
                "dm_ratio_pct",
                "dm_abs_difference_pp",
                "dm_relative_difference",
                "dm_issue",
                "dm_pct_used",
                "KgMS/ha",
                "biomass_kg_ha_recomputed",
                "biomass_recompute_difference",
                "kgms_workbook_status",
                "dm_pct_status",
            ]
        ].merge(
            quality[
                key
                + [
                    "REG. DE LAB.",
                    "% N",
                    "n_pct_primary",
                    "% FDA ",
                    "% FDN ",
                    "KgN/ha",
                    "INN",
                    "Origen del dato",
                    "quality_status",
                    "n_pct_cell_status",
                    "adf_cell_status",
                    "ndf_cell_status",
                    "q_workbook_cell_status",
                    "nni_workbook_cell_status",
                ]
            ],
            on=key,
            how="left",
            validate="one_to_one",
        )
        longitudinal = longitudinal.rename(
            columns={
                "Fecha": "date",
                "Muestra": "sample_id",
                "Condición": "sector",
                "Tratamiento": "treatment",
                "Repetición": "block",
                "%MS": "dm_pct_recorded",
                "KgMS/ha": "biomass_kg_ha_workbook",
                "biomass_kg_ha_recomputed": "biomass_kg_ha",
                "% N": "n_pct_recorded",
                "n_pct_primary": "n_pct",
                "% FDA ": "adf_pct",
                "% FDN ": "ndf_pct",
                "KgN/ha": "q_kg_n_ha_workbook",
                "INN": "nni_workbook",
                "REG. DE LAB.": "lab_id",
                "Origen del dato": "data_origin",
            }
        )
        longitudinal["q_kg_n_ha"] = (
            longitudinal["biomass_kg_ha"] * longitudinal["n_pct"] / 100.0
        )
        biomass_t_ha = longitudinal["biomass_kg_ha"] / 1000.0
        valid_biomass = biomass_t_ha.where(biomass_t_ha.gt(0.0))
        critical_primary = nni_primary_coefficient * valid_biomass.pow(
            nni_primary_exponent
        )
        critical_sensitivity = nni_sensitivity_coefficient * valid_biomass.pow(
            nni_sensitivity_exponent
        )
        longitudinal["nni_primary"] = longitudinal["n_pct"] / critical_primary
        longitudinal["nni_sensitivity"] = longitudinal["n_pct"] / critical_sensitivity
        longitudinal["q_workbook_recompute_difference"] = (
            longitudinal["q_kg_n_ha"] - longitudinal["q_kg_n_ha_workbook"]
        )
        longitudinal["nni_sensitivity_workbook_difference"] = (
            longitudinal["nni_sensitivity"] - longitudinal["nni_workbook"]
        )
        longitudinal["plot_id"] = (
            longitudinal["sector"].astype(str)
            + "_"
            + longitudinal["block"].astype(str)
            + "_"
            + longitudinal["treatment"].astype(str)
        )

        sample_dates = tuple(
            pd.Timestamp(value)
            for value in sorted(longitudinal["date"].dropna().unique())
        )
        if not sample_dates:
            raise ValueError("No se encontraron fechas experimentales en Datos_MS.")
        date_labels = {
            value: pd.Timestamp(value).strftime("%d %b").lower()
            for value in sample_dates
        }
        longitudinal["date_label"] = longitudinal["date"].map(date_labels)

        # Harvest rows are selected by treatment before positional renaming to
        # avoid footer text in the worksheet.
        harvest = raw_harvest.copy()
        harvest["_excel_row"] = np.arange(6, 6 + len(harvest))
        harvest["Condición"] = harvest["Condición"].map(_normalize_sector)
        harvest["Tratamiento"] = harvest["Tratamiento"].map(_normalize_treatment)
        harvest["Repetición"] = harvest["Repetición"].map(_normalize_block)
        harvest = harvest.loc[
            harvest["Tratamiento"].str.fullmatch(TREATMENT_PATTERN, na=False)
        ].copy()
        columns = list(harvest.columns)
        rename_by_position = {
            columns[0]: "date",
            columns[1]: "sample_id",
            columns[2]: "sector",
            columns[3]: "treatment",
            columns[4]: "block",
            columns[5]: "panicle_count",
            columns[6]: "dirty_mass_g",
            columns[7]: "clean_mass_g",
            columns[8]: "w100_1_g",
            columns[9]: "w100_2_g",
            columns[10]: "w100_3_g",
            columns[11]: "w1000_workbook_g",
        }
        harvest = harvest.rename(columns=rename_by_position)
        keep = list(rename_by_position.values()) + ["_excel_row"]
        harvest = harvest[keep].copy()
        harvest["date"] = pd.to_datetime(harvest["date"], errors="coerce")
        harvest["sample_id"] = pd.to_numeric(
            harvest["sample_id"], errors="coerce"
        ).astype("Int64")
        harvest = _coerce_numeric(
            harvest,
            [
                "panicle_count",
                "dirty_mass_g",
                "clean_mass_g",
                "w100_1_g",
                "w100_2_g",
                "w100_3_g",
                "w1000_workbook_g",
            ],
        )
        harvest["w1000_g"] = (
            harvest[["w100_1_g", "w100_2_g", "w100_3_g"]].mean(axis=1) * 10.0
        )
        harvest["w1000_workbook_status"] = [
            _cell_status(
                formula_coordinates=formula_coordinates,
                sheet="Datos_Rto",
                column_number=12,
                excel_row=int(row),
                semantically_derived=True,
            )
            for row in harvest["_excel_row"]
        ]
        harvest["panicle_density_m2"] = (
            harvest["panicle_count"] / harvest_sample_area_m2
        )
        harvest["dirty_yield_kg_ha"] = (
            harvest["dirty_mass_g"] * 10.0 / harvest_sample_area_m2
        )
        harvest["clean_yield_kg_ha"] = (
            harvest["clean_mass_g"] * 10.0 / harvest_sample_area_m2
        )
        harvest["clean_recovery"] = harvest["clean_mass_g"] / harvest["dirty_mass_g"]
        harvest["cleaning_loss_pct"] = 100.0 * (1.0 - harvest["clean_recovery"])
        harvest["estimated_seed_count"] = (
            1000.0 * harvest["clean_mass_g"] / harvest["w1000_g"]
        )
        harvest["estimated_seeds_per_panicle"] = (
            harvest["estimated_seed_count"] / harvest["panicle_count"]
        )
        harvest["plot_id"] = (
            harvest["sector"].astype(str)
            + "_"
            + harvest["block"].astype(str)
            + "_"
            + harvest["treatment"].astype(str)
        )

        final_date = max(sample_dates)
        final_biomass = longitudinal.loc[
            longitudinal["date"].eq(final_date),
            ["plot_id", "biomass_kg_ha", "dm_issue"],
        ]
        harvest = harvest.merge(
            final_biomass, on="plot_id", how="left", validate="one_to_one"
        )
        harvest["harvest_index_pct"] = (
            100.0 * harvest["clean_yield_kg_ha"] / harvest["biomass_kg_ha"]
        )

        application_dates = [
            pd.Timestamp(cast(Any, value))
            for value in schedule[["first_application", "second_application"]]
            .stack()
            .dropna()
        ]
        study_start = min([*application_dates, *sample_dates])
        study_end = max(max(sample_dates), pd.Timestamp(harvest["date"].max()))
        management, water_monthly, water_period_totals = _parse_management_and_water(
            value_book,
            experiment_year=experiment_year,
            study_start=pd.Timestamp(study_start),
            study_end=pd.Timestamp(study_end),
        )
        water_map = water_period_totals.set_index("sector")["gross_input_mm"]
        harvest["gross_water_input_mm"] = harvest["sector"].map(water_map)
        harvest["apparent_water_productivity"] = (
            harvest["clean_yield_kg_ha"] / harvest["gross_water_input_mm"]
        )

        m0_reference = harvest.loc[
            harvest["treatment"].eq("M0"),
            ["sector", "block", "clean_yield_kg_ha"],
        ].rename(columns={"clean_yield_kg_ha": "m0_yield_same_block"})
        harvest = harvest.merge(
            m0_reference, on=["sector", "block"], how="left", validate="many_to_one"
        )
        harvest["agronomic_efficiency"] = np.where(
            harvest["treatment"].eq("M0"),
            np.nan,
            (harvest["clean_yield_kg_ha"] - harvest["m0_yield_same_block"])
            / experimental_n_total,
        )

        treatments = tuple(schedule["treatment"].tolist())
        sectors = tuple(
            sorted(
                longitudinal["sector"].dropna().unique(),
                key=lambda value: (
                    ["Secano", "Riego"].index(value)
                    if value in {"Secano", "Riego"}
                    else 99
                ),
            )
        )
        blocks = tuple(sorted(longitudinal["block"].dropna().unique()))
        schedule["extra_n_kg_ha"] = np.where(
            schedule["treatment"].eq("M0"), 0.0, experimental_n_total
        )
        schedule["dose_per_application_kg_ha"] = np.where(
            schedule["treatment"].eq("M0"),
            0.0,
            dose_per_application,
        )
        schedule["application_count"] = np.where(
            schedule["treatment"].eq("M0"), 0, application_count
        )

        audit = _source_audit(
            value_sheet=value_book,
            schedule=schedule,
            sample_dates=sample_dates,
            experiment_year=experiment_year,
            formula_coordinates=formula_coordinates,
        )
        spec = ExperimentSpec(
            workbook_path=path,
            source_sha256=sha256_file(path),
            experiment_year=experiment_year,
            treatments=treatments,
            sectors=sectors,
            blocks=blocks,
            repetitions=repetitions,
            plot_area_m2=plot_area_m2,
            row_spacing_m=row_spacing_m,
            biomass_sample_area_m2=biomass_sample_area_m2,
            harvest_sample_area_m2=harvest_sample_area_m2,
            experimental_n_total_kg_ha=experimental_n_total,
            applications_per_treatment=application_count,
            dose_per_application_kg_ha=dose_per_application,
            schedule=schedule.reset_index(drop=True),
            management=management.reset_index(drop=True),
            water_monthly=water_monthly.reset_index(drop=True),
            water_period_totals=water_period_totals.reset_index(drop=True),
            source_audit=audit.reset_index(drop=True),
        )

        for frame in (longitudinal, harvest):
            frame["sector"] = pd.Categorical(
                frame["sector"], categories=list(sectors), ordered=True
            )
            frame["block"] = pd.Categorical(
                frame["block"], categories=list(blocks), ordered=True
            )
            frame["treatment"] = pd.Categorical(
                frame["treatment"], categories=list(treatments), ordered=True
            )
        longitudinal["date"] = pd.Categorical(
            longitudinal["date"], categories=list(sample_dates), ordered=True
        )

        seed_weight_long = harvest.melt(
            id_vars=["plot_id", "sector", "block", "treatment", "sample_id"],
            value_vars=["w100_1_g", "w100_2_g", "w100_3_g"],
            var_name="technical_replicate",
            value_name="w100_g",
        )

        baseline_biomass = baseline_ms.loc[
            baseline_ms["biomass_kg_ha_recomputed"].notna(),
            ["Fecha", "Muestra", "Condición", "Repetición", "biomass_kg_ha_recomputed"],
        ].rename(
            columns={
                "Fecha": "date",
                "Muestra": "sample_id",
                "Condición": "sector",
                "Repetición": "block",
                "biomass_kg_ha_recomputed": "biomass_kg_ha",
            }
        )
        baseline_tillers = baseline_ms.loc[
            baseline_ms["Macollos/m2"].notna(),
            ["Fecha", "Muestra", "Condición", "Repetición", "Macollos/m2"],
        ].rename(
            columns={
                "Fecha": "date",
                "Muestra": "sample_id",
                "Condición": "sector",
                "Repetición": "replicate_label",
                "Macollos/m2": "tillers_m2",
            }
        )

        longitudinal = longitudinal.sort_values(
            ["sector", "block", "treatment", "date"]
        ).reset_index(drop=True)
        harvest = harvest.sort_values(["sector", "block", "treatment"]).reset_index(
            drop=True
        )
        qa = _build_qa(
            spec=spec,
            longitudinal=longitudinal,
            harvest=harvest,
            sample_dates=sample_dates,
        )
        failed_errors = qa.loc[qa["severity"].eq("error") & ~qa["passes"]]
        if not failed_errors.empty:
            raise AssertionError(failed_errors.to_string(index=False))

        return ExperimentData(
            spec=spec,
            longitudinal=longitudinal,
            harvest=harvest,
            seed_weight_long=seed_weight_long.reset_index(drop=True),
            baseline_biomass=baseline_biomass.reset_index(drop=True),
            baseline_tillers=baseline_tillers.reset_index(drop=True),
            variable_lineage=_variable_lineage(),
            qa=qa,
        )
    finally:
        value_book.close()


def source_provenance_table(data: ExperimentData) -> pd.DataFrame:
    """Compact provenance table suitable for notebook display."""

    spec = data.spec
    return pd.DataFrame(
        [
            {
                "source_file": spec.workbook_path.name,
                "source_path": str(spec.workbook_path),
                "sha256": spec.source_sha256,
                "experiment_year": spec.experiment_year,
                "dry_matter_rows": len(data.longitudinal),
                "harvest_rows": len(data.harvest),
            }
        ]
    )
