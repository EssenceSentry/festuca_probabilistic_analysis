"""Rebuild a coherent Excel workbook from the canonical CSV/JSON bundle."""

# openpyxl has incomplete type information at several public API boundaries.
# pyright: reportArgumentType=false, reportAssignmentType=false
# pyright: reportAttributeAccessIssue=false, reportCallIssue=false
# pyright: reportGeneralTypeIssues=false, reportUnknownArgumentType=false
# pyright: reportUnknownMemberType=false, reportUnknownVariableType=false

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Final, Literal, cast

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.table import (
    Table,
    TableColumn,
    TableStyleInfo,
)
from openpyxl.worksheet.worksheet import Worksheet

from festuca_analysis.source_data import (
    DEFAULT_DATA_DIR,
    REQUIRED_FORMULA_TARGETS,
    locate_data_dir,
    sha256_data_bundle,
    validate_data_bundle,
)

PROJECT_ROOT: Final = Path(__file__).resolve().parents[2]
DEFAULT_OUTPUT_PATH: Final = PROJECT_ROOT / "dist" / "datos_festuca_canonicos.xlsx"

HEADER_ROW: Final = 4
FIRST_DATA_ROW: Final = HEADER_ROW + 1
DOCUMENT_ROLE: Final = "documental"
RECORDED_ROLE: Final = "registrado"
CALCULATED_ROLE: Final = "calculado"

SHEET_ORDER: Final = (
    "Índice",
    "Diccionario",
    "Fórmulas",
    "Diseño",
    "Metadatos",
    "Cronograma",
    "Suelo",
    "Manejo",
    "Agua",
    "MS registrada",
    "MS calculada",
    "Cosecha registrada",
    "Cosecha calculada",
    "Calidad registrada",
    "Calidad calculada",
    "Calidad faltante",
)

DATASET_LAYOUT: Final = {
    "experimental_design.csv": ("Diseño", "tblDiseno"),
    "experiment_metadata.csv": ("Metadatos", "tblMetadatos"),
    "field_timeline.csv": ("Cronograma", "tblCronograma"),
    "soil_analysis.csv": ("Suelo", "tblSuelo"),
    "field_management.csv": ("Manejo", "tblManejo"),
    "water_inputs.csv": ("Agua", "tblAgua"),
    "dry_matter_recorded.csv": ("MS registrada", "tblMSRegistrada"),
    "dry_matter_calculated.csv": ("MS calculada", "tblMSCalculada"),
    "harvest_recorded.csv": ("Cosecha registrada", "tblCosechaRegistrada"),
    "harvest_calculated.csv": ("Cosecha calculada", "tblCosechaCalculada"),
    "quality_recorded.csv": ("Calidad registrada", "tblCalidadRegistrada"),
    "quality_calculated.csv": ("Calidad calculada", "tblCalidadCalculada"),
    "missing_quality_estimate.csv": ("Calidad faltante", "tblCalidadFaltante"),
}

DOCUMENT_TABLES: Final = {
    "Índice": "tblIndice",
    "Diccionario": "tblDiccionario",
    "Fórmulas": "tblFormulas",
}

CALCULATED_SOURCES: Final = {
    "dry_matter_calculated.csv": "dry_matter_recorded.csv",
    "harvest_calculated.csv": "harvest_recorded.csv",
    "quality_calculated.csv": "quality_recorded.csv",
}

CALCULATED_HELPERS: Final = {
    "harvest_calculated.csv": (
        ("final_biomass_kg_ha", "Biomasa final enlazada", "number", "kg MS/ha"),
    ),
    "quality_calculated.csv": (
        (
            "biomass_reported_kg_ha",
            "Biomasa informada enlazada",
            "number",
            "kg MS/ha",
        ),
        ("effective_n_pct", "Nitrógeno efectivo", "number", "%"),
    ),
}

SUPPORTED_FORMULA_IDS: Final = frozenset(
    {
        "dry_matter_pct_from_weights",
        "dry_matter_abs_difference_pp",
        "dry_matter_relative_difference",
        "dry_matter_issue",
        "biomass_recorded_dm",
        "biomass_ratio_dm",
        "biomass_recalculation_difference",
        "tiller_density",
        "thousand_seed_weight",
        "panicle_density",
        "dirty_yield",
        "clean_yield",
        "clean_recovery",
        "cleaning_loss",
        "estimated_seed_count",
        "estimated_seeds_per_panicle",
        "harvest_index",
        "rcbd_missing_n",
        "rcbd_missing_adf",
        "rcbd_missing_ndf",
        "quality_n_estimate_lookup",
        "quality_adf_estimate_lookup",
        "quality_ndf_estimate_lookup",
        "n_accumulated",
        "critical_n",
        "nni",
        "missing_n_accumulated",
        "missing_critical_n",
        "missing_nni",
    }
)

ROLE_COLORS: Final = {
    DOCUMENT_ROLE: "5B6573",
    RECORDED_ROLE: "1F4E78",
    CALCULATED_ROLE: "548235",
}
ROLE_LIGHT_FILLS: Final = {
    DOCUMENT_ROLE: "E7E9EC",
    RECORDED_ROLE: "D9EAF7",
    CALCULATED_ROLE: "E2F0D9",
}
WARNING_FILL: Final = PatternFill("solid", fgColor="FFF2CC")
ESTIMATE_FILL: Final = PatternFill("solid", fgColor="FCE4D6")
DATE_FORMAT: Final = "yyyy-mm-dd"
NUMBER_FORMAT: Final = "0.0000"
INTEGER_FORMAT: Final = "0"

ColumnRole = Literal["static", "lookup", "formula", "helper"]


@dataclass(frozen=True)
class ColumnDefinition:
    """A visible workbook column and its canonical meaning."""

    technical_name: str
    header: str
    data_type: str
    unit: str | None
    nullable: bool
    origin_file: str
    role: ColumnRole


@dataclass(frozen=True)
class SheetDefinition:
    """Workbook location and table metadata for one canonical dataset."""

    filename: str
    sheet_name: str
    table_name: str
    role: str
    grain: str
    primary_key: tuple[str, ...]
    columns: tuple[ColumnDefinition, ...]


def _read_object(path: Path) -> dict[str, object]:
    payload: object = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise TypeError(f"Se esperaba un objeto JSON: {path}")
    return cast(dict[str, object], payload)


def _manifest_datasets(
    manifest: Mapping[str, object],
) -> dict[str, dict[str, object]]:
    raw = manifest.get("datasets")
    if not isinstance(raw, dict):
        raise TypeError("manifest.json debe contener el objeto datasets")
    result: dict[str, dict[str, object]] = {}
    for filename, spec in raw.items():
        if not isinstance(filename, str) or not isinstance(spec, dict):
            raise TypeError("Cada dataset del manifiesto debe ser un objeto")
        result[filename] = cast(dict[str, object], spec)
    return result


def _column_specs(dataset: Mapping[str, object]) -> dict[str, dict[str, object]]:
    raw = dataset.get("columns")
    if not isinstance(raw, dict):
        raise TypeError("Cada dataset debe declarar columns")
    result: dict[str, dict[str, object]] = {}
    for name, spec in raw.items():
        if not isinstance(name, str) or not isinstance(spec, dict):
            raise TypeError("Cada columna del manifiesto debe ser un objeto")
        result[name] = cast(dict[str, object], spec)
    return result


def _header(column_spec: Mapping[str, object]) -> str:
    name = column_spec.get("display_name_es")
    if not isinstance(name, str) or not name:
        raise TypeError("Cada columna debe tener display_name_es")
    unit = column_spec.get("unit")
    return f"{name} ({unit})" if isinstance(unit, str) and unit else name


def _canonical_column(
    filename: str,
    technical_name: str,
    spec: Mapping[str, object],
    *,
    role: ColumnRole,
) -> ColumnDefinition:
    data_type = spec.get("type")
    if not isinstance(data_type, str):
        raise TypeError(f"Tipo inválido en {filename}:{technical_name}")
    unit = spec.get("unit")
    return ColumnDefinition(
        technical_name=technical_name,
        header=_header(spec),
        data_type=data_type,
        unit=unit if isinstance(unit, str) and unit else None,
        nullable=bool(spec.get("nullable", False)),
        origin_file=filename,
        role=role,
    )


def _build_sheet_definitions(
    datasets: Mapping[str, Mapping[str, object]],
) -> dict[str, SheetDefinition]:
    if set(datasets) != set(DATASET_LAYOUT):
        missing = sorted(set(DATASET_LAYOUT) - set(datasets))
        unexpected = sorted(set(datasets) - set(DATASET_LAYOUT))
        raise ValueError(
            f"Datasets incompatibles con el reconstructor; missing={missing}, "
            f"unexpected={unexpected}"
        )

    definitions: dict[str, SheetDefinition] = {}
    for filename, dataset in datasets.items():
        columns = _column_specs(dataset)
        dataset_role = dataset.get("role")
        if dataset_role not in {"recorded", "calculated"}:
            raise ValueError(f"Rol no soportado en {filename}: {dataset_role!r}")
        role = RECORDED_ROLE if dataset_role == "recorded" else CALCULATED_ROLE
        visible: list[ColumnDefinition] = []
        source_file = CALCULATED_SOURCES.get(filename)
        if source_file is not None:
            source_columns = _column_specs(datasets[source_file])
            primary_key_raw = dataset.get("primary_key")
            if not isinstance(primary_key_raw, list):
                raise TypeError(f"Clave primaria inválida en {filename}")
            primary_key = {str(value) for value in primary_key_raw}
            for name, spec in source_columns.items():
                column_role: ColumnRole = "static" if name in primary_key else "lookup"
                visible.append(
                    _canonical_column(source_file, name, spec, role=column_role)
                )
            for name, display, data_type, unit in CALCULATED_HELPERS.get(filename, ()):
                visible.append(
                    ColumnDefinition(
                        technical_name=name,
                        header=f"{display} ({unit})" if unit else display,
                        data_type=data_type,
                        unit=unit,
                        nullable=False,
                        origin_file="vista_calculada",
                        role="helper",
                    )
                )
            for name, spec in columns.items():
                if name in primary_key:
                    continue
                formula_role: ColumnRole = (
                    "formula"
                    if (filename, name) in REQUIRED_FORMULA_TARGETS
                    else "static"
                )
                visible.append(
                    _canonical_column(filename, name, spec, role=formula_role)
                )
        else:
            for name, spec in columns.items():
                formula_role = (
                    "formula"
                    if (filename, name) in REQUIRED_FORMULA_TARGETS
                    else "static"
                )
                visible.append(
                    _canonical_column(filename, name, spec, role=formula_role)
                )

        headers = [column.header for column in visible]
        if len(headers) != len(set(headers)):
            raise ValueError(f"Encabezados duplicados en la vista de {filename}")
        primary_key_raw = dataset.get("primary_key")
        if not isinstance(primary_key_raw, list) or not all(
            isinstance(value, str) for value in primary_key_raw
        ):
            raise TypeError(f"Clave primaria inválida en {filename}")
        grain = dataset.get("grain_es")
        if not isinstance(grain, str):
            raise TypeError(f"Descripción de grano inválida en {filename}")
        sheet_name, table_name = DATASET_LAYOUT[filename]
        definitions[filename] = SheetDefinition(
            filename=filename,
            sheet_name=sheet_name,
            table_name=table_name,
            role=role,
            grain=grain,
            primary_key=tuple(cast(list[str], primary_key_raw)),
            columns=tuple(visible),
        )
    return definitions


def _formula_definitions(
    payload: Mapping[str, object],
) -> list[dict[str, object]]:
    raw = payload.get("formulas")
    if not isinstance(raw, list):
        raise TypeError("formulas.json debe contener una lista formulas")
    formulas: list[dict[str, object]] = []
    identifiers: set[str] = set()
    targets: set[tuple[str, str]] = set()
    for item in raw:
        if not isinstance(item, dict):
            raise TypeError("Cada definición de fórmula debe ser un objeto")
        formula = cast(dict[str, object], item)
        identifier = formula.get("id")
        output_file = formula.get("output_file")
        output_column = formula.get("output_column")
        if not all(
            isinstance(value, str) for value in (identifier, output_file, output_column)
        ):
            raise TypeError(
                "Toda fórmula debe declarar id, output_file y output_column"
            )
        identifier_text = cast(str, identifier)
        target = (cast(str, output_file), cast(str, output_column))
        if identifier_text not in SUPPORTED_FORMULA_IDS:
            raise ValueError(f"Fórmula desconocida: {identifier_text}")
        if identifier_text in identifiers or target in targets:
            raise ValueError(f"Fórmula o destino duplicado: {identifier_text}")
        identifiers.add(identifier_text)
        targets.add(target)
        formulas.append(formula)
    if identifiers != set(SUPPORTED_FORMULA_IDS):
        missing = sorted(set(SUPPORTED_FORMULA_IDS) - identifiers)
        raise ValueError(f"Faltan fórmulas soportadas: {missing}")
    if targets != set(REQUIRED_FORMULA_TARGETS):
        raise ValueError("Los destinos de formulas.json no coinciden con el esquema")
    return formulas


def _formula_map(
    formulas: Sequence[Mapping[str, object]],
) -> dict[tuple[str, str], Mapping[str, object]]:
    return {
        (
            cast(str, formula["output_file"]),
            cast(str, formula["output_column"]),
        ): formula
        for formula in formulas
    }


def _structured_row(table_name: str, header: str) -> str:
    return f"{table_name}[[#This Row],[{header.replace(']', ']]')}]]"


def _structured_column(table_name: str, header: str) -> str:
    return f"{table_name}[{header.replace(']', ']]')}]"


def _index_match(
    lookup_value: str,
    table_name: str,
    key_header: str,
    return_header: str,
) -> str:
    return (
        f"INDEX({_structured_column(table_name, return_header)},"
        f"MATCH({lookup_value},"
        f"{_structured_column(table_name, key_header)},0))"
    )


def _column_map(definition: SheetDefinition) -> dict[str, ColumnDefinition]:
    return {column.technical_name: column for column in definition.columns}


def _compile_formula(
    formula: Mapping[str, object],
    definition: SheetDefinition,
    definitions: Mapping[str, SheetDefinition],
    metadata_parameters: set[str],
) -> str:
    identifier = formula.get("id")
    if not isinstance(identifier, str) or identifier not in SUPPORTED_FORMULA_IDS:
        raise ValueError(f"Fórmula desconocida: {identifier!r}")
    template = formula.get("excel_formula_template")
    inputs = formula.get("inputs")
    if not isinstance(template, str) or not template.startswith("="):
        raise TypeError(f"Plantilla Excel inválida en {identifier}")
    if not isinstance(inputs, list) or not all(
        isinstance(value, str) for value in inputs
    ):
        raise TypeError(f"Entradas inválidas en {identifier}")

    columns = _column_map(definition)
    missing_definition = definitions["missing_quality_estimate.csv"]
    missing_columns = _column_map(missing_definition)
    external_missing = {
        "n_pct_estimated",
        "adf_pct_estimated",
        "ndf_pct_estimated",
    }
    unresolved = [
        value
        for value in cast(list[str], inputs)
        if value not in columns
        and value not in metadata_parameters
        and not (
            definition.filename == "quality_calculated.csv"
            and value in external_missing
        )
    ]
    if unresolved:
        raise ValueError(f"Entradas no resolubles en {identifier}: {unresolved}")

    compiled = template
    xlookup_pattern = re.compile(r"XLOOKUP\(([^,]+),([^,]+),([^)]+)\)")

    def replace_xlookup(match: re.Match[str]) -> str:
        lookup_value, lookup_range, return_range = match.groups()
        return f"INDEX({return_range},MATCH({lookup_value},{lookup_range},0))"

    compiled = xlookup_pattern.sub(replace_xlookup, compiled)
    if "XLOOKUP(" in compiled:
        raise ValueError(f"XLOOKUP no pudo compilarse en {identifier}: {compiled}")

    range_pattern = re.compile(r"\[@([A-Za-z0-9_]+)\]:\[@([A-Za-z0-9_]+)\]")

    def replace_range(match: re.Match[str]) -> str:
        first, last = match.groups()
        if first not in columns or last not in columns:
            raise ValueError(f"Rango no resoluble en {identifier}: {first}:{last}")
        first_header = columns[first].header.replace("]", "]]")
        last_header = columns[last].header.replace("]", "]]")
        return (
            f"{definition.table_name}[[#This Row]," f"[{first_header}]:[{last_header}]]"
        )

    compiled = range_pattern.sub(replace_range, compiled)
    for name in sorted(columns, key=len, reverse=True):
        compiled = compiled.replace(
            f"[@{name}]",
            _structured_row(definition.table_name, columns[name].header),
        )

    for name in external_missing:
        if name in missing_columns:
            compiled = compiled.replace(
                f"MissingQuality[{name}]",
                _structured_column(
                    missing_definition.table_name, missing_columns[name].header
                ),
            )
    missing_key = missing_columns["sample_id"]
    compiled = compiled.replace(
        "MissingQuality[sample_id]",
        _structured_column(missing_definition.table_name, missing_key.header),
    )

    metadata_definition = definitions["experiment_metadata.csv"]
    metadata_columns = _column_map(metadata_definition)
    for parameter in sorted(metadata_parameters, key=len, reverse=True):
        if not re.search(
            rf"(?<![A-Za-z0-9_]){re.escape(parameter)}(?![A-Za-z0-9_])", compiled
        ):
            continue
        lookup = _index_match(
            f'"{parameter}"',
            metadata_definition.table_name,
            metadata_columns["parameter"].header,
            metadata_columns["value"].header,
        )
        compiled = re.sub(
            rf"(?<![A-Za-z0-9_]){re.escape(parameter)}(?![A-Za-z0-9_])",
            lookup,
            compiled,
        )

    for local_name in ("treatment_count", "block_count"):
        if local_name in columns:
            compiled = re.sub(
                rf"(?<![A-Za-z0-9_]){local_name}(?![A-Za-z0-9_])",
                _structured_row(definition.table_name, columns[local_name].header),
                compiled,
            )

    if "MissingQuality[" in compiled or re.search(r"\[@[A-Za-z0-9_]+\]", compiled):
        raise ValueError(
            f"Plantilla no compilada completamente en {identifier}: {compiled}"
        )
    return compiled


def _native_value(value: object) -> object:
    if value is None or value is pd.NA or pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, np.generic):
        return value.item()
    return value


def _metadata_native_value(row: pd.Series) -> object:
    value = _native_value(row["value"])
    value_type = str(row["value_type"])
    if value is None:
        return None
    if value_type == "integer":
        return int(str(value))
    if value_type == "number":
        return float(str(value))
    if value_type == "boolean":
        return str(value).casefold() == "true"
    return str(value)


def _number_format(column: ColumnDefinition) -> str:
    if column.data_type == "date":
        return DATE_FORMAT
    if column.data_type == "integer":
        return INTEGER_FORMAT
    if column.data_type == "number":
        return NUMBER_FORMAT
    return "General"


def _initialize_sheet(
    worksheet: Worksheet,
    *,
    title: str,
    subtitle: str,
    role: str,
) -> None:
    color = ROLE_COLORS[role]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = f"B{FIRST_DATA_ROW}"
    worksheet["A1"] = title
    worksheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    worksheet["A1"].fill = PatternFill("solid", fgColor=color)
    worksheet["A1"].alignment = Alignment(vertical="center")
    worksheet["A2"] = subtitle
    worksheet["A2"].font = Font(size=10, italic=True, color="404040")
    worksheet["A2"].fill = PatternFill("solid", fgColor=ROLE_LIGHT_FILLS[role])
    worksheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.row_dimensions[1].height = 25
    worksheet.row_dimensions[2].height = 31


def _write_table(
    worksheet: Worksheet,
    *,
    headers: Sequence[str],
    rows: Sequence[Sequence[object]],
    table_name: str,
    role: str,
    column_definitions: Sequence[ColumnDefinition] | None = None,
    preserve_formula_text: bool = False,
) -> None:
    if not headers or not rows:
        raise ValueError(f"La tabla {table_name} debe tener encabezados y filas")
    if len(headers) != len(set(headers)):
        raise ValueError(f"Encabezados duplicados en {table_name}")
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=HEADER_ROW, column=column_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=ROLE_COLORS[role])
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_index, row in enumerate(rows, start=FIRST_DATA_ROW):
        if len(row) != len(headers):
            raise ValueError(f"Fila con longitud inválida en {table_name}")
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            if (
                preserve_formula_text
                and isinstance(value, str)
                and value.startswith("=")
            ):
                cell.data_type = "s"
            cell.alignment = Alignment(vertical="top")
            if column_definitions is not None:
                cell.number_format = _number_format(
                    column_definitions[column_index - 1]
                )
    last_row = HEADER_ROW + len(rows)
    last_column = len(headers)
    for column_index in range(1, last_column + 1):
        worksheet.cell(row=1, column=column_index).fill = PatternFill(
            "solid", fgColor=ROLE_COLORS[role]
        )
        worksheet.cell(row=2, column=column_index).fill = PatternFill(
            "solid", fgColor=ROLE_LIGHT_FILLS[role]
        )
    table_columns: list[TableColumn] = []
    for column_index, header in enumerate(headers):
        values = [row[column_index] for row in rows]
        live_formulas = [
            value
            for value in values
            if isinstance(value, str) and value.startswith("=")
        ]
        if (
            live_formulas
            and not preserve_formula_text
            and (len(live_formulas) != len(values) or len(set(live_formulas)) != 1)
        ):
            raise ValueError(
                f"La columna calculada {table_name}:{header} no tiene una fórmula uniforme"
            )
        table_columns.append(
            TableColumn(
                id=column_index + 1,
                name=header,
            )
        )
    table = Table(
        displayName=table_name,
        ref=f"A{HEADER_ROW}:{get_column_letter(last_column)}{last_row}",
        totalsRowShown=False,
        tableColumns=table_columns,
    )
    table.tableStyleInfo = TableStyleInfo(
        name=("TableStyleMedium2" if role == RECORDED_ROLE else "TableStyleMedium4"),
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)
    worksheet.row_dimensions[HEADER_ROW].height = 38
    for column_index, header in enumerate(headers, start=1):
        values = [header]
        for row in rows[:75]:
            value = row[column_index - 1]
            if value is not None and not (
                isinstance(value, str) and value.startswith("=")
            ):
                values.append(str(value))
        longest = max(len(value) for value in values)
        width = min(max(longest + 2, 11), 42)
        if any(
            token in header.casefold()
            for token in ("descripción", "actividad", "expresión", "plantilla")
        ):
            width = 42
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _source_lookup_formula(
    definition: SheetDefinition,
    column: ColumnDefinition,
    definitions: Mapping[str, SheetDefinition],
) -> str:
    source_file = CALCULATED_SOURCES[definition.filename]
    source = definitions[source_file]
    source_columns = _column_map(source)
    if len(definition.primary_key) != 1:
        raise ValueError(f"La vista {definition.filename} requiere clave simple")
    key = definition.primary_key[0]
    current_columns = _column_map(definition)
    lookup_value = _structured_row(definition.table_name, current_columns[key].header)
    lookup = _index_match(
        lookup_value,
        source.table_name,
        source_columns[key].header,
        source_columns[column.technical_name].header,
    )
    if not column.nullable:
        return "=" + lookup
    populated_count = (
        f"COUNTIFS({_structured_column(source.table_name, source_columns[key].header)},"
        f"{lookup_value},"
        f"{_structured_column(source.table_name, source_columns[column.technical_name].header)},"
        '"<>")'
    )
    return f'=IF({populated_count}=0,"",{lookup})'


def _helper_formula(
    filename: str,
    technical_name: str,
    definition: SheetDefinition,
    definitions: Mapping[str, SheetDefinition],
) -> str:
    columns = _column_map(definition)
    dry_definition = definitions["dry_matter_calculated.csv"]
    dry_columns = _column_map(dry_definition)
    if filename == "harvest_calculated.csv" and technical_name == "final_biomass_kg_ha":
        observation = _structured_row(
            definition.table_name, columns["observation_id"].header
        )
        lookup_value = f'SUBSTITUTE({observation},"harvest_","dm_")'
        return "=" + _index_match(
            lookup_value,
            dry_definition.table_name,
            dry_columns["observation_id"].header,
            dry_columns["biomass_recorded_dm_kg_ha"].header,
        )
    if (
        filename == "quality_calculated.csv"
        and technical_name == "biomass_reported_kg_ha"
    ):
        return "=" + _index_match(
            _structured_row(definition.table_name, columns["sample_id"].header),
            dry_definition.table_name,
            dry_columns["sample_id"].header,
            dry_columns["biomass_reported_kg_ha"].header,
        )
    if filename == "quality_calculated.csv" and technical_name == "effective_n_pct":
        recorded = _structured_row(definition.table_name, columns["n_pct"].header)
        estimated = _structured_row(
            definition.table_name, columns["estimated_n_pct"].header
        )
        return f'=IF({recorded}="",{estimated},{recorded})'
    raise ValueError(f"Ayudante desconocido: {filename}:{technical_name}")


def _calculated_rows(
    definition: SheetDefinition,
    frame: pd.DataFrame,
    definitions: Mapping[str, SheetDefinition],
    formula_lookup: Mapping[tuple[str, str], Mapping[str, object]],
    metadata_parameters: set[str],
) -> list[list[object]]:
    rows: list[list[object]] = []
    for _, record in frame.iterrows():
        row: list[object] = []
        for column in definition.columns:
            if column.role == "lookup":
                row.append(_source_lookup_formula(definition, column, definitions))
            elif column.role == "helper":
                row.append(
                    _helper_formula(
                        definition.filename,
                        column.technical_name,
                        definition,
                        definitions,
                    )
                )
            elif column.role == "formula":
                target = (definition.filename, column.technical_name)
                formula = formula_lookup.get(target)
                if formula is None:
                    raise ValueError(f"Falta fórmula declarada para {target}")
                row.append(
                    _compile_formula(
                        formula, definition, definitions, metadata_parameters
                    )
                )
            else:
                row.append(_native_value(record[column.technical_name]))
        rows.append(row)
    return rows


def _recorded_rows(
    definition: SheetDefinition, frame: pd.DataFrame
) -> list[list[object]]:
    rows: list[list[object]] = []
    for _, record in frame.iterrows():
        row: list[object] = []
        for column in definition.columns:
            if (
                definition.filename == "experiment_metadata.csv"
                and column.technical_name == "value"
            ):
                row.append(_metadata_native_value(record))
            else:
                row.append(_native_value(record[column.technical_name]))
        rows.append(row)
    return rows


def _apply_conditional_formatting(
    worksheet: Worksheet, definition: SheetDefinition, row_count: int
) -> None:
    if row_count <= 0:
        return
    headers = {
        column.technical_name: index
        for index, column in enumerate(definition.columns, 1)
    }
    last_column = get_column_letter(len(definition.columns))
    area = f"A{FIRST_DATA_ROW}:{last_column}{HEADER_ROW + row_count}"
    if definition.filename == "dry_matter_calculated.csv":
        issue_column = get_column_letter(headers["dry_matter_issue"])
        worksheet.conditional_formatting.add(
            area,
            FormulaRule(
                formula=[f"${issue_column}{FIRST_DATA_ROW}=TRUE"],
                fill=WARNING_FILL,
            ),
        )
    if definition.filename == "quality_recorded.csv":
        status_column = get_column_letter(headers["measurement_status"])
        worksheet.conditional_formatting.add(
            area,
            FormulaRule(
                formula=[f'${status_column}{FIRST_DATA_ROW}="estimated"'],
                fill=ESTIMATE_FILL,
            ),
        )
    if definition.filename == "quality_calculated.csv":
        status_column = get_column_letter(headers["measurement_status"])
        worksheet.conditional_formatting.add(
            area,
            FormulaRule(
                formula=[f'${status_column}{FIRST_DATA_ROW}="estimated"'],
                fill=ESTIMATE_FILL,
            ),
        )
    if definition.filename == "missing_quality_estimate.csv":
        worksheet.conditional_formatting.add(
            area,
            FormulaRule(formula=["TRUE"], fill=ESTIMATE_FILL),
        )


def _relation_text(dataset: Mapping[str, object], column_name: str) -> str:
    relations: list[str] = []
    raw = dataset.get("foreign_keys", [])
    if not isinstance(raw, list):
        raise TypeError("foreign_keys debe ser una lista")
    for item in raw:
        if not isinstance(item, dict):
            raise TypeError("Cada relación debe ser un objeto")
        relation = cast(dict[str, object], item)
        columns = relation.get("columns")
        referenced = relation.get("references")
        reference_columns = relation.get("reference_columns", columns)
        if (
            isinstance(columns, list)
            and column_name in columns
            and isinstance(referenced, str)
            and isinstance(reference_columns, list)
        ):
            position = columns.index(column_name)
            relations.append(f"{referenced}:{reference_columns[position]}")
    return "; ".join(relations)


def _document_rows(
    datasets: Mapping[str, Mapping[str, object]],
    definitions: Mapping[str, SheetDefinition],
    formulas_payload: Mapping[str, object],
    bundle_hash: str,
    frames: Mapping[str, pd.DataFrame],
) -> dict[str, tuple[list[str], list[list[object]]]]:
    index_headers = [
        "Hoja",
        "CSV de origen",
        "Rol",
        "Filas",
        "Clave primaria",
        "SHA-256 del conjunto",
    ]
    index_rows: list[list[object]] = [
        [
            definitions[filename].sheet_name,
            filename,
            definitions[filename].role,
            len(frames[filename]),
            " + ".join(definitions[filename].primary_key),
            bundle_hash,
        ]
        for filename in DATASET_LAYOUT
    ]

    dictionary_headers = [
        "Hoja",
        "CSV de origen",
        "Nombre técnico",
        "Nombre español",
        "Tipo",
        "Unidad",
        "Admite vacío",
        "Relación",
    ]
    dictionary_rows: list[list[object]] = []
    for filename in DATASET_LAYOUT:
        dataset = datasets[filename]
        sheet_name = definitions[filename].sheet_name
        for technical_name, spec in _column_specs(dataset).items():
            dictionary_rows.append(
                [
                    sheet_name,
                    filename,
                    technical_name,
                    cast(str, spec["display_name_es"]),
                    cast(str, spec["type"]),
                    spec.get("unit"),
                    bool(spec.get("nullable", False)),
                    _relation_text(dataset, technical_name),
                ]
            )
    for filename, helpers in CALCULATED_HELPERS.items():
        for technical_name, display, data_type, unit in helpers:
            dictionary_rows.append(
                [
                    definitions[filename].sheet_name,
                    "vista_calculada",
                    technical_name,
                    display,
                    data_type,
                    unit,
                    False,
                    "Derivada mediante búsqueda entre tablas",
                ]
            )

    raw_order = formulas_payload.get("evaluation_order")
    if not isinstance(raw_order, list) or not all(
        isinstance(value, str) for value in raw_order
    ):
        raise TypeError("evaluation_order debe ser una lista de cadenas")
    evaluation_order = cast(list[str], raw_order)
    stage_for_file = {
        "dry_matter_calculated.csv": "dry_matter_calculations",
        "missing_quality_estimate.csv": "missing_quality_estimate",
        "quality_calculated.csv": "quality_calculations",
        "harvest_calculated.csv": "harvest_calculations",
    }
    formulas = _formula_definitions(formulas_payload)
    formula_headers = [
        "Orden",
        "Etapa",
        "Identificador",
        "CSV de salida",
        "Columna de salida",
        "Entradas",
        "Expresión semántica",
        "Unidad",
        "Plantilla Excel",
    ]
    formula_rows: list[list[object]] = []
    for ordinal, formula in enumerate(formulas, start=1):
        output_file = cast(str, formula["output_file"])
        stage = stage_for_file[output_file]
        inputs = cast(list[str], formula["inputs"])
        formula_rows.append(
            [
                ordinal,
                f"{evaluation_order.index(stage) + 1}. {stage}",
                formula["id"],
                output_file,
                formula["output_column"],
                ", ".join(inputs),
                formula["expression"],
                formula.get("unit"),
                formula["excel_formula_template"],
            ]
        )
    return {
        "Índice": (index_headers, index_rows),
        "Diccionario": (dictionary_headers, dictionary_rows),
        "Fórmulas": (formula_headers, formula_rows),
    }


def _build_workbook(data_dir: Path) -> Workbook:
    frames = validate_data_bundle(data_dir)
    manifest = _read_object(data_dir / "manifest.json")
    formulas_payload = _read_object(data_dir / "formulas.json")
    datasets = _manifest_datasets(manifest)
    definitions = _build_sheet_definitions(datasets)
    formulas = _formula_definitions(formulas_payload)
    formula_lookup = _formula_map(formulas)
    bundle_hash = sha256_data_bundle(data_dir)
    metadata_parameters = set(
        frames["experiment_metadata.csv"]["parameter"].astype(str)
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation = CalcProperties(
        calcId=0,
        calcMode="auto",
        fullCalcOnLoad=True,
        forceFullCalc=True,
    )
    workbook.properties.creator = "festuca-rebuild-workbook"
    workbook.properties.title = "Datos canónicos del ensayo de festuca"
    workbook.properties.subject = "Artefacto derivado de data/"
    fixed_time = datetime(2025, 1, 1, tzinfo=UTC)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time

    for sheet_name in SHEET_ORDER:
        workbook.create_sheet(sheet_name)

    document_rows = _document_rows(
        datasets, definitions, formulas_payload, bundle_hash, frames
    )
    for sheet_name in ("Índice", "Diccionario", "Fórmulas"):
        worksheet = workbook[sheet_name]
        headers, rows = document_rows[sheet_name]
        _initialize_sheet(
            worksheet,
            title=sheet_name,
            subtitle=(
                "Documentación generada desde manifest.json y formulas.json; "
                f"SHA-256 del conjunto: {bundle_hash}"
            ),
            role=DOCUMENT_ROLE,
        )
        _write_table(
            worksheet,
            headers=headers,
            rows=rows,
            table_name=DOCUMENT_TABLES[sheet_name],
            role=DOCUMENT_ROLE,
            preserve_formula_text=sheet_name == "Fórmulas",
        )

    # Populate the dependency table before formulas that refer to it. Tab order stays fixed.
    build_order = [
        filename for filename in DATASET_LAYOUT if filename != "quality_calculated.csv"
    ]
    build_order.append("quality_calculated.csv")
    for filename in build_order:
        definition = definitions[filename]
        worksheet = workbook[definition.sheet_name]
        _initialize_sheet(
            worksheet,
            title=definition.sheet_name,
            subtitle=(
                f"{definition.grain} Origen canónico: {filename}. "
                f"Rol: {definition.role}."
            ),
            role=definition.role,
        )
        if definition.role == CALCULATED_ROLE:
            rows = _calculated_rows(
                definition,
                frames[filename],
                definitions,
                formula_lookup,
                metadata_parameters,
            )
        else:
            rows = _recorded_rows(definition, frames[filename])
        _write_table(
            worksheet,
            headers=[column.header for column in definition.columns],
            rows=rows,
            table_name=definition.table_name,
            role=definition.role,
            column_definitions=definition.columns,
        )
        _apply_conditional_formatting(worksheet, definition, len(rows))

    if workbook.sheetnames != list(SHEET_ORDER):
        raise AssertionError("El orden de hojas cambió durante la construcción")
    return workbook


def _validate_rebuilt_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=False, read_only=False)
    if workbook.sheetnames != list(SHEET_ORDER):
        raise ValueError(f"Orden de hojas inesperado: {workbook.sheetnames}")
    table_names: list[str] = []
    for worksheet in workbook.worksheets:
        tables = list(worksheet.tables.values())
        if len(tables) != 1:
            raise ValueError(
                f"{worksheet.title} debe contener exactamente una tabla, tiene {len(tables)}"
            )
        table_names.append(tables[0].displayName)
    if len(table_names) != len(set(table_names)):
        raise ValueError("Los nombres de tabla no son únicos")
    for sheet_name in (
        "Diseño",
        "Metadatos",
        "Cronograma",
        "Suelo",
        "Manejo",
        "Agua",
        "MS registrada",
        "Cosecha registrada",
        "Calidad registrada",
    ):
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=FIRST_DATA_ROW):
            if any(cell.data_type == "f" for cell in row):
                raise ValueError(f"La hoja registrada {sheet_name} contiene fórmulas")
    formula_count = 0
    for sheet_name in (
        "MS calculada",
        "Cosecha calculada",
        "Calidad calculada",
        "Calidad faltante",
    ):
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=FIRST_DATA_ROW):
            for cell in row:
                if cell.data_type == "f":
                    formula_count += 1
                    formula_text = str(cell.value)
                    if any(
                        error in formula_text
                        for error in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                    ):
                        raise ValueError(
                            f"Error literal en fórmula {sheet_name}!{cell.coordinate}"
                        )
    if formula_count == 0:
        raise ValueError("Las vistas calculadas no contienen fórmulas vivas")
    if workbook.calculation.calcMode != "auto":
        raise ValueError("El libro no está configurado para cálculo automático")
    workbook.close()


def _cell_signature(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _semantic_signature(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    sheets: list[dict[str, object]] = []
    for worksheet in workbook.worksheets:
        tables = list(worksheet.tables.values())
        cells = [
            [
                (
                    _cell_signature(cell.value),
                    cell.data_type,
                    cell.number_format,
                )
                for cell in row
            ]
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            )
        ]
        widths = {
            key: dimension.width
            for key, dimension in worksheet.column_dimensions.items()
        }
        sheets.append(
            {
                "name": worksheet.title,
                "cells": cells,
                "tables": [
                    (
                        table.displayName,
                        table.ref,
                        table.tableStyleInfo.name if table.tableStyleInfo else None,
                    )
                    for table in tables
                ],
                "freeze": str(worksheet.freeze_panes),
                "gridlines": worksheet.sheet_view.showGridLines,
                "conditional_formats": len(worksheet.conditional_formatting),
                "widths": widths,
            }
        )
    signature: dict[str, object] = {
        "sheets": sheets,
        "calculation": (
            workbook.calculation.calcMode,
            workbook.calculation.fullCalcOnLoad,
            workbook.calculation.forceFullCalc,
        ),
    }
    workbook.close()
    return signature


def rebuild_workbook(
    data_dir: Path | str,
    output_path: Path | str,
    *,
    check: bool = False,
    force: bool = False,
) -> Path:
    """Build or semantically check the derived canonical XLSX workbook."""

    if check and force:
        raise ValueError("--check y --force son mutuamente excluyentes")
    canonical_dir = locate_data_dir(data_dir)
    output = Path(output_path).expanduser().resolve()
    if output.suffix.casefold() != ".xlsx":
        raise ValueError(f"La salida debe ser un archivo .xlsx: {output}")
    lock_file = output.with_name(f"~${output.name}")
    if lock_file.exists():
        raise RuntimeError(
            f"Close Microsoft Excel before rebuilding the workbook: {lock_file}"
        )
    validate_data_bundle(canonical_dir)
    if check and not output.is_file():
        raise FileNotFoundError(f"No existe el libro para comprobar: {output}")
    if not check and output.exists() and not force:
        raise FileExistsError(
            f"La salida ya existe; use --force para reemplazarla: {output}"
        )

    output.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output.stem}.", suffix=".xlsx", dir=output.parent
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook = _build_workbook(canonical_dir)
        workbook.save(temporary_path)
        workbook.close()
        _validate_rebuilt_workbook(temporary_path)
        if check:
            if _semantic_signature(output) != _semantic_signature(temporary_path):
                raise ValueError(
                    f"El libro existente no coincide semánticamente con data/: {output}"
                )
            return output
        os.replace(temporary_path, output)
        return output
    finally:
        temporary_path.unlink(missing_ok=True)


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Reconstruye el XLSX derivado desde los CSV/JSON canónicos."
    )
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--force", action="store_true")
    mode.add_argument("--check", action="store_true")
    return parser


def main() -> None:
    """Command-line entry point."""

    args = _parser().parse_args()
    output = rebuild_workbook(
        args.data_dir,
        args.output,
        check=bool(args.check),
        force=bool(args.force),
    )
    action = "verificado" if args.check else "reconstruido"
    print(f"XLSX {action}: {output}")


if __name__ == "__main__":
    main()
