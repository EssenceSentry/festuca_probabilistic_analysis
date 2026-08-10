"""One-time migration from the historical workbook to canonical CSV data."""

from __future__ import annotations

# openpyxl's workbook boundary is only partially typed.
# pyright: reportUnknownArgumentType=false, reportUnknownMemberType=false
# pyright: reportUnknownVariableType=false
import argparse
import csv
import hashlib
import json
import math
import os
import re
import shutil
import tempfile
import unicodedata
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Final, cast

from openpyxl import load_workbook

PROJECT_ROOT: Final = Path(__file__).resolve().parents[2]
DEFAULT_SOURCE: Final = PROJECT_ROOT / "sources" / "Datos_Ema_Serrana_INN.xlsx"
DEFAULT_OUTPUT_DIR: Final = PROJECT_ROOT / "data"
SCHEMA_VERSION: Final = 1
CANONICAL_DBCA_ESTIMATES: Final = {
    "n_pct": 2.8688484862162937,
    "adf_pct": 40.77623836505554,
    "ndf_pct": 69.57708690222144,
}


@dataclass(frozen=True)
class Arguments:
    source: Path
    output_dir: Path
    check: bool
    force: bool


def parse_args() -> Arguments:
    parser = argparse.ArgumentParser(
        description=(
            "Migrate the historical Festuca workbook to normalized canonical CSVs."
        )
    )
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--check",
        action="store_true",
        help="Compare the canonical directory with a fresh workbook migration.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Replace an existing canonical directory after a successful migration.",
    )
    namespace = parser.parse_args()
    return Arguments(
        source=cast(Path, namespace.source),
        output_dir=cast(Path, namespace.output_dir),
        check=cast(bool, namespace.check),
        force=cast(bool, namespace.force),
    )


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(
        character for character in normalized if not unicodedata.combining(character)
    )


def _text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _treatment(value: object) -> str:
    return _text(value).upper().replace("MO", "M0")


def _sector(value: object) -> str:
    normalized = _strip_accents(_text(value)).casefold()
    return {"secano": "Secano", "riego": "Riego"}.get(normalized, _text(value))


def _block(value: object) -> str:
    return _text(value).upper()


def _iso_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raise TypeError(f"Expected an Excel date, received {value!r}")


def _number(value: object) -> float | None:
    if value is None or value == "":
        return None
    number = float(cast(Any, value))
    return None if math.isnan(number) else number


def _integer(value: object) -> int | None:
    number = _number(value)
    return None if number is None else int(number)


def _safe_divide(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator is None or denominator == 0.0:
        return None
    return numerator / denominator


def _csv_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _write_csv(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    if not rows:
        raise ValueError(f"Refusing to create an empty canonical dataset: {path.name}")
    fieldnames = list(rows[0])
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            if list(row) != fieldnames:
                raise ValueError(f"Inconsistent columns while writing {path.name}")
            writer.writerow({key: _csv_value(value) for key, value in row.items()})


def _write_json(path: Path, payload: object) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def _column(
    display_name_es: str,
    data_type: str,
    *,
    unit: str | None = None,
    nullable: bool = False,
    description_es: str | None = None,
) -> dict[str, object]:
    result: dict[str, object] = {
        "display_name_es": display_name_es,
        "type": data_type,
        "nullable": nullable,
    }
    if unit is not None:
        result["unit"] = unit
    if description_es is not None:
        result["description_es"] = description_es
    return result


def _metadata_rows(book: Any) -> list[dict[str, object]]:
    ensayo = book["Ensayo"]
    design = book["Diseño"]
    row_spacing_cm = float(book["Datos_MS"]["K3"].value)
    n_total = float(ensayo["B8"].value)
    applications = int(ensayo["B9"].value)
    source_description = _text(book["Datos_Rto"]["D3"].value)
    harvest_match = re.search(
        r"\(\s*(\d+(?:[.,]\d+)?)\s*metro[s]?\s*[,;x]\s*(\d+)\s*surcos?\s*\)",
        _strip_accents(source_description).casefold(),
    )
    if harvest_match is None:
        raise ValueError(
            "No se pudo interpretar la geometría de cosecha en Datos_Rto!D3."
        )
    harvest_length_per_row_m = float(harvest_match.group(1).replace(",", "."))
    harvest_row_count = int(harvest_match.group(2))
    row_spacing_m = row_spacing_cm / 100.0
    records: list[tuple[str, object, str, str | None, str]] = [
        ("producer", ensayo["B3"].value, "string", None, "Ensayo!B3"),
        ("forage_common_name", "Festuca", "string", None, "Ensayo!A4"),
        ("cultivar", ensayo["B4"].value, "string", None, "Ensayo!B4"),
        ("sowing_year", ensayo["B5"].value, "integer", "year", "Ensayo!B5"),
        ("experiment_year", ensayo["B6"].value, "integer", "year", "Ensayo!B6"),
        (
            "experimental_n_total_kg_ha",
            n_total,
            "number",
            "kg N/ha",
            "Ensayo!B8",
        ),
        (
            "applications_per_treatment",
            applications,
            "integer",
            "applications",
            "Ensayo!B9",
        ),
        (
            "dose_per_application_kg_ha",
            n_total / applications,
            "number",
            "kg N/ha",
            "derived from Ensayo!B8:B9",
        ),
        ("plot_area_m2", ensayo["B10"].value, "number", "m2", "Ensayo!B10"),
        ("plot_width_m", ensayo["B11"].value, "number", "m", "Ensayo!B11"),
        ("plot_length_m", ensayo["B12"].value, "number", "m", "Ensayo!B12"),
        ("repetitions", ensayo["B13"].value, "integer", "blocks", "Ensayo!B13"),
        (
            "urea_per_plot_kg",
            ensayo["B14"].value,
            "number",
            "kg product/plot",
            "Ensayo!B14",
        ),
        (
            "urea_rate_kg_ha",
            ensayo["B15"].value,
            "number",
            "kg product/ha",
            "Ensayo!B15",
        ),
        ("urea_n_pct", 46, "number", "% N", "Ensayo!A15"),
        (
            "row_spacing_cm",
            row_spacing_cm,
            "number",
            "cm",
            "Datos_MS!K3",
        ),
        ("biomass_cut_length_m", 1, "number", "m", "Datos_MS!G5"),
        (
            "biomass_sample_area_m2",
            row_spacing_m,
            "number",
            "m2",
            "derived from Datos_MS!K3 and G5",
        ),
        (
            "harvest_length_per_row_m",
            harvest_length_per_row_m,
            "number",
            "m",
            "Datos_Rto!D3",
        ),
        (
            "harvest_row_count",
            harvest_row_count,
            "integer",
            "rows",
            "Datos_Rto!D3",
        ),
        (
            "harvest_sample_area_m2",
            harvest_length_per_row_m * harvest_row_count * row_spacing_m,
            "number",
            "m2",
            "derived from Datos_Rto!D3 and Datos_MS!K3",
        ),
        (
            "stakes_per_trial",
            design["F1"].value,
            "integer",
            "stakes",
            "Diseño!F1:G1",
        ),
        ("north_marker", design["A6"].value, "string", None, "Diseño!A6"),
        (
            "reference_boundary",
            design["A1"].value,
            "string",
            None,
            "Diseño!A1",
        ),
    ]
    return [
        {
            "parameter": parameter,
            "value": value,
            "value_type": value_type,
            "unit": unit,
            "legacy_source": source,
        }
        for parameter, value, value_type, unit, source in records
    ]


def _design_rows(book: Any) -> list[dict[str, object]]:
    sheet = book["Diseño"]
    rows: list[dict[str, object]] = []
    for block_number, excel_row in enumerate((6, 13, 20, 27), start=1):
        for position, column in enumerate(range(3, 9), start=1):
            rows.append(
                {
                    "block": f"R{block_number}",
                    "plot_position": position,
                    "treatment": _treatment(sheet.cell(excel_row, column).value),
                }
            )
    return rows


def _timeline_rows() -> list[dict[str, object]]:
    events: list[tuple[str, str, str, str, str | None, int | None, str | None, str]] = [
        (
            "2025-06-12_trial_installation",
            "2025-06-12",
            "trial_installation",
            "whole_trial",
            None,
            None,
            None,
            "Instalación del ensayo.",
        ),
        (
            "2025-06-12_tiller_count",
            "2025-06-12",
            "tiller_count",
            "baseline",
            None,
            None,
            None,
            "Conteo inicial de macollos.",
        ),
        (
            "2025-06-12_biomass_quality_sampling",
            "2025-06-12",
            "biomass_quality_sampling",
            "baseline",
            None,
            None,
            None,
            "Muestreo inicial de biomasa y calidad.",
        ),
        (
            "2025-06-12_m1_application_1",
            "2025-06-12",
            "nitrogen_application",
            "treatment",
            "M1",
            1,
            None,
            "Primera aplicación de M1.",
        ),
        (
            "2025-06-27_m2_application_1",
            "2025-06-27",
            "nitrogen_application",
            "treatment",
            "M2",
            1,
            None,
            "Primera aplicación de M2.",
        ),
        (
            "2025-07-09_m3_application_1",
            "2025-07-09",
            "nitrogen_application",
            "treatment",
            "M3",
            1,
            None,
            "Primera aplicación de M3.",
        ),
        (
            "2025-08-04_m1_application_2",
            "2025-08-04",
            "nitrogen_application",
            "treatment",
            "M1",
            2,
            None,
            "Segunda aplicación de M1.",
        ),
        (
            "2025-08-04_m2_application_2",
            "2025-08-04",
            "nitrogen_application",
            "treatment",
            "M2",
            2,
            None,
            "Segunda aplicación de M2.",
        ),
        (
            "2025-08-04_m4_application_1",
            "2025-08-04",
            "nitrogen_application",
            "treatment",
            "M4",
            1,
            None,
            "Primera aplicación de M4.",
        ),
        (
            "2025-08-21_m3_application_2",
            "2025-08-21",
            "nitrogen_application",
            "treatment",
            "M3",
            2,
            None,
            "Segunda aplicación de M3.",
        ),
        (
            "2025-08-25_m5_application_1",
            "2025-08-25",
            "nitrogen_application",
            "treatment",
            "M5",
            1,
            None,
            "Primera aplicación de M5.",
        ),
        (
            "2025-09-16_m4_application_2",
            "2025-09-16",
            "nitrogen_application",
            "treatment",
            "M4",
            2,
            None,
            "Segunda aplicación de M4; el orden intradía no está documentado.",
        ),
        (
            "2025-09-16_m5_application_2",
            "2025-09-16",
            "nitrogen_application",
            "treatment",
            "M5",
            2,
            None,
            "Segunda aplicación de M5; el orden intradía no está documentado.",
        ),
        (
            "2025-09-16_biomass_quality_sampling",
            "2025-09-16",
            "biomass_quality_sampling",
            "all_treatments",
            None,
            None,
            "Z31; registro adicional Z3.08",
            "Primer corte de materia seca y muestreo de calidad; el orden intradía no está documentado.",
        ),
        (
            "2025-09-16_phenology_observation",
            "2025-09-16",
            "phenology_observation",
            "all_treatments",
            None,
            None,
            "Z31; registro adicional Z3.08",
            "Nudos registrados: 0-1-0-1-1-1-1-2-1-0.",
        ),
        (
            "2025-10-20_biomass_quality_sampling",
            "2025-10-20",
            "biomass_quality_sampling",
            "all_treatments",
            None,
            None,
            "anthesis",
            "Corte para materia seca y calidad en antesis.",
        ),
        (
            "2025-11-12_biomass_quality_sampling",
            "2025-11-12",
            "biomass_quality_sampling",
            "all_treatments",
            None,
            None,
            None,
            "Corte final para materia seca y calidad.",
        ),
        (
            "2025-11-12_harvest",
            "2025-11-12",
            "harvest",
            "all_treatments",
            None,
            None,
            None,
            "Cosecha para evaluar el rendimiento y sus componentes.",
        ),
    ]
    ambiguous_dates = {"2025-06-12", "2025-08-04", "2025-09-16", "2025-11-12"}
    return [
        {
            "event_id": event_id,
            "date": event_date,
            "event_type": event_type,
            "scope": scope,
            "treatment": treatment,
            "application_number": application_number,
            "growth_stage": growth_stage,
            "description": description,
            "within_day_order_known": (
                False if event_date in ambiguous_dates else None
            ),
        }
        for (
            event_id,
            event_date,
            event_type,
            scope,
            treatment,
            application_number,
            growth_stage,
            description,
        ) in events
    ]


def _soil_rows(book: Any) -> list[dict[str, object]]:
    sheet = book["Manejo"]
    parameters = [
        (2, "phosphorus", "ppm"),
        (3, "potassium", "meq"),
        (4, "base_saturation", "%"),
        (5, "organic_matter", "%"),
        (6, "ph", None),
        (7, "calcium", "meq"),
        (8, "magnesium", "meq"),
        (9, "sodium", None),
    ]
    return [
        {
            "parameter": parameter,
            "value": _number(sheet.cell(row, 2).value),
            "unit": unit,
        }
        for row, parameter, unit in parameters
    ]


def _management_rows(book: Any) -> list[dict[str, object]]:
    sheet = book["Manejo"]
    return [
        {
            "year": 2025,
            "month": row - 13,
            "month_label": _text(sheet.cell(row, 1).value),
            "activity": _text(sheet.cell(row, 2).value) or None,
        }
        for row in range(14, 25)
    ]


def _water_rows(book: Any) -> list[dict[str, object]]:
    sheet = book["Manejo"]
    return [
        {
            "year": 2025,
            "month": row - 13,
            "month_label": _text(sheet.cell(row, 6).value),
            "irrigation_mm": _number(sheet.cell(row, 7).value) or 0.0,
            "rainfall_mm": _number(sheet.cell(row, 8).value) or 0.0,
        }
        for row in range(14, 25)
    ]


def _observation_id(
    prefix: str, event_date: str, sector: str, treatment: str, block: str
) -> str:
    parts = [prefix, event_date.replace("-", ""), sector, treatment, block]
    return "_".join(_strip_accents(part).casefold() for part in parts)


def _dry_matter_rows(
    book: Any, metadata: Mapping[str, object]
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    sheet = book["Datos_MS"]
    area = float(cast(Any, metadata["biomass_sample_area_m2"]))
    row_spacing_cm = float(cast(Any, metadata["row_spacing_cm"]))
    recorded: list[dict[str, object]] = []
    calculated: list[dict[str, object]] = []
    for row in range(6, 160):
        raw_date = sheet.cell(row, 1).value
        if raw_date is None:
            continue
        event_date = _iso_date(raw_date)
        sector = _sector(sheet.cell(row, 3).value)
        treatment = _treatment(sheet.cell(row, 4).value)
        block = _block(sheet.cell(row, 5).value)
        observation_id = _observation_id("dm", event_date, sector, treatment, block)
        tillers = _number(sheet.cell(row, 6).value)
        green_1m = _number(sheet.cell(row, 7).value)
        green_sample = _number(sheet.cell(row, 8).value)
        tray = _number(sheet.cell(row, 9).value)
        dry_weight = _number(sheet.cell(row, 10).value)
        dm_pct = _number(sheet.cell(row, 11).value)
        biomass_reported = _number(sheet.cell(row, 12).value)
        ratio_fraction = _safe_divide(dry_weight, green_sample)
        ratio_pct = None if ratio_fraction is None else 100.0 * ratio_fraction
        absolute_difference = (
            None if dm_pct is None or ratio_pct is None else abs(dm_pct - ratio_pct)
        )
        relative_difference = (
            None
            if absolute_difference is None or dm_pct in (None, 0.0)
            else absolute_difference / abs(dm_pct)
        )
        dm_issue = bool(
            absolute_difference is not None
            and relative_difference is not None
            and absolute_difference >= 5.0
            and relative_difference >= 0.20
        )
        biomass_recorded = (
            None
            if green_1m is None or dm_pct is None
            else green_1m * (dm_pct / 100.0) * 10.0 / area
        )
        biomass_ratio = (
            None
            if green_1m is None or ratio_pct is None
            else green_1m * (ratio_pct / 100.0) * 10.0 / area
        )
        tiller_density = (
            None
            if tillers is None
            else tillers * (100.0 / 30.0) * (100.0 / row_spacing_cm)
        )
        recorded.append(
            {
                "observation_id": observation_id,
                "date": event_date,
                "sample_id": _integer(sheet.cell(row, 2).value),
                "sector": sector,
                "treatment": treatment,
                "block": block,
                "tillers_30_cm": tillers,
                "green_weight_1m_g": green_1m,
                "green_weight_sample_g": green_sample,
                "tray_weight_g": tray,
                "dry_weight_g": dry_weight,
                "dry_matter_pct": dm_pct,
            }
        )
        calculated.append(
            {
                "observation_id": observation_id,
                "dry_matter_pct_from_weights": ratio_pct,
                "dry_matter_abs_difference_pp": absolute_difference,
                "dry_matter_relative_difference": relative_difference,
                "dry_matter_issue": dm_issue,
                "biomass_reported_kg_ha": biomass_reported,
                "biomass_recorded_dm_kg_ha": biomass_recorded,
                "biomass_ratio_dm_kg_ha": biomass_ratio,
                "biomass_recalculation_difference_kg_ha": (
                    None
                    if biomass_reported is None or biomass_recorded is None
                    else biomass_recorded - biomass_reported
                ),
                "tiller_density_m2": tiller_density,
            }
        )
    return recorded, calculated


def _harvest_rows(
    book: Any,
    metadata: Mapping[str, object],
    biomass_by_plot: Mapping[tuple[str, str, str], float],
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    sheet = book["Datos_Rto"]
    area = float(cast(Any, metadata["harvest_sample_area_m2"]))
    recorded: list[dict[str, object]] = []
    calculated: list[dict[str, object]] = []
    for row in range(6, 54):
        event_date = _iso_date(sheet.cell(row, 1).value)
        sector = _sector(sheet.cell(row, 3).value)
        treatment = _treatment(sheet.cell(row, 4).value)
        block = _block(sheet.cell(row, 5).value)
        observation_id = _observation_id(
            "harvest", event_date, sector, treatment, block
        )
        panicles = cast(float, _number(sheet.cell(row, 6).value))
        dirty_mass = cast(float, _number(sheet.cell(row, 7).value))
        clean_mass = cast(float, _number(sheet.cell(row, 8).value))
        replicates = [
            cast(float, _number(sheet.cell(row, column).value))
            for column in range(9, 12)
        ]
        w1000 = sum(replicates) / len(replicates) * 10.0
        clean_recovery = clean_mass / dirty_mass
        estimated_seed_count = 1000.0 * clean_mass / w1000
        biomass = biomass_by_plot[(sector, treatment, block)]
        clean_yield = clean_mass * 10.0 / area
        recorded.append(
            {
                "observation_id": observation_id,
                "date": event_date,
                "sample_id": _integer(sheet.cell(row, 2).value),
                "sector": sector,
                "treatment": treatment,
                "block": block,
                "panicle_count": panicles,
                "dirty_seed_mass_g": dirty_mass,
                "clean_seed_mass_g": clean_mass,
                "w100_rep1_g": replicates[0],
                "w100_rep2_g": replicates[1],
                "w100_rep3_g": replicates[2],
            }
        )
        calculated.append(
            {
                "observation_id": observation_id,
                "w1000_g": w1000,
                "panicle_density_m2": panicles / area,
                "dirty_yield_kg_ha": dirty_mass * 10.0 / area,
                "clean_yield_kg_ha": clean_yield,
                "clean_recovery": clean_recovery,
                "cleaning_loss_pct": 100.0 * (1.0 - clean_recovery),
                "estimated_seed_count": estimated_seed_count,
                "estimated_seeds_per_panicle": estimated_seed_count / panicles,
                "harvest_index_pct": 100.0 * clean_yield / biomass,
            }
        )
    return recorded, calculated


def _rcbd_components(
    rows: Sequence[Mapping[str, object]],
    *,
    value_column: str,
    missing_treatment: str,
    missing_block: str,
    treatments: int,
    blocks: int,
) -> tuple[float, float, float, float]:
    observed = [row for row in rows if row[value_column] is not None]
    block_total = sum(
        float(cast(Any, row[value_column]))
        for row in observed
        if row["block"] == missing_block
    )
    treatment_total = sum(
        float(cast(Any, row[value_column]))
        for row in observed
        if row["treatment"] == missing_treatment
    )
    grand_total = sum(float(cast(Any, row[value_column])) for row in observed)
    estimate = (treatments * block_total + blocks * treatment_total - grand_total) / (
        (treatments - 1) * (blocks - 1)
    )
    return block_total, treatment_total, grand_total, estimate


def _quality_rows(
    book: Any,
    biomass_by_sample: Mapping[int, float],
) -> tuple[
    list[dict[str, object]],
    list[dict[str, object]],
    list[dict[str, object]],
]:
    sheet = book["Calidad"]
    recorded: list[dict[str, object]] = []
    for row in range(2, 154):
        sample_id = cast(int, _integer(sheet.cell(row, 2).value))
        origin = _strip_accents(_text(sheet.cell(row, 14).value)).casefold()
        status = "estimated" if "estim" in origin else "measured"
        recorded.append(
            {
                "sample_id": sample_id,
                "date": _iso_date(sheet.cell(row, 1).value),
                "sector": _sector(sheet.cell(row, 3).value),
                "treatment": _treatment(sheet.cell(row, 4).value),
                "block": _block(sheet.cell(row, 5).value),
                "lab_registration": _integer(sheet.cell(row, 8).value),
                "n_pct": (
                    None if status == "estimated" else _number(sheet.cell(row, 9).value)
                ),
                "adf_pct": (
                    None
                    if status == "estimated"
                    else _number(sheet.cell(row, 10).value)
                ),
                "ndf_pct": (
                    None
                    if status == "estimated"
                    else _number(sheet.cell(row, 11).value)
                ),
                "measurement_status": status,
                "data_origin": _text(sheet.cell(row, 14).value),
            }
        )

    missing = [row for row in recorded if row["measurement_status"] == "estimated"]
    if len(missing) != 1:
        raise ValueError(f"Expected one estimated quality row, found {len(missing)}")
    missing_row = missing[0]
    same_design = [
        row
        for row in recorded
        if row["date"] == missing_row["date"] and row["sector"] == missing_row["sector"]
    ]
    components: dict[str, tuple[float, float, float, float]] = {}
    for variable in ("n_pct", "adf_pct", "ndf_pct"):
        calculated_components = _rcbd_components(
            same_design,
            value_column=variable,
            missing_treatment=cast(str, missing_row["treatment"]),
            missing_block=cast(str, missing_row["block"]),
            treatments=6,
            blocks=4,
        )
        canonical_estimate = CANONICAL_DBCA_ESTIMATES[variable]
        if not math.isclose(
            calculated_components[3], canonical_estimate, rel_tol=1e-15, abs_tol=1e-15
        ):
            raise ValueError(
                f"DBCA estimate drift for {variable}: "
                f"{calculated_components[3]} != {canonical_estimate}"
            )
        components[variable] = (*calculated_components[:3], canonical_estimate)
    sample_id = cast(int, missing_row["sample_id"])
    biomass = biomass_by_sample[sample_id]
    critical_n = 4.8 * (biomass / 1000.0) ** -0.32
    n_estimate = components["n_pct"][3]
    estimate_row = {
        "sample_id": sample_id,
        "date": missing_row["date"],
        "sector": missing_row["sector"],
        "treatment": missing_row["treatment"],
        "block": missing_row["block"],
        "biomass_kg_ha": biomass,
        "treatment_count": 6,
        "block_count": 4,
        "n_block_total": components["n_pct"][0],
        "n_treatment_total": components["n_pct"][1],
        "n_grand_total": components["n_pct"][2],
        "n_pct_estimated": n_estimate,
        "adf_block_total": components["adf_pct"][0],
        "adf_treatment_total": components["adf_pct"][1],
        "adf_grand_total": components["adf_pct"][2],
        "adf_pct_estimated": components["adf_pct"][3],
        "ndf_block_total": components["ndf_pct"][0],
        "ndf_treatment_total": components["ndf_pct"][1],
        "ndf_grand_total": components["ndf_pct"][2],
        "ndf_pct_estimated": components["ndf_pct"][3],
        "n_accumulated_kg_ha": biomass * n_estimate / 100.0,
        "critical_n_pct": critical_n,
        "nni": n_estimate / critical_n,
    }
    calculated: list[dict[str, object]] = []
    for row in recorded:
        row_sample_id = cast(int, row["sample_id"])
        if row["measurement_status"] == "estimated":
            n_value = n_estimate
            estimated_n = n_estimate
            estimated_adf = components["adf_pct"][3]
            estimated_ndf = components["ndf_pct"][3]
            status = "from_dbca_estimate"
        else:
            n_value = cast(float, row["n_pct"])
            estimated_n = None
            estimated_adf = None
            estimated_ndf = None
            status = "from_measured_quality"
        row_biomass = biomass_by_sample[row_sample_id]
        row_critical = 4.8 * (row_biomass / 1000.0) ** -0.32
        calculated.append(
            {
                "sample_id": row_sample_id,
                "estimated_n_pct": estimated_n,
                "estimated_adf_pct": estimated_adf,
                "estimated_ndf_pct": estimated_ndf,
                "n_accumulated_kg_ha": row_biomass * n_value / 100.0,
                "critical_n_pct": row_critical,
                "nni": n_value / row_critical,
                "calculation_status": status,
            }
        )
    return recorded, calculated, [estimate_row]


def _formula_payload() -> dict[str, object]:
    def formula(
        identifier: str,
        output_file: str,
        output_column: str,
        inputs: Sequence[str],
        expression: str,
        excel_template: str,
        unit: str | None,
    ) -> dict[str, object]:
        return {
            "id": identifier,
            "output_file": output_file,
            "output_column": output_column,
            "inputs": list(inputs),
            "expression": expression,
            "excel_formula_template": excel_template,
            "unit": unit,
        }

    formulas = [
        formula(
            "dry_matter_pct_from_weights",
            "dry_matter_calculated.csv",
            "dry_matter_pct_from_weights",
            ["dry_weight_g", "green_weight_sample_g"],
            "100 * dry_weight_g / green_weight_sample_g",
            "=100*[@dry_weight_g]/[@green_weight_sample_g]",
            "%",
        ),
        formula(
            "dry_matter_abs_difference_pp",
            "dry_matter_calculated.csv",
            "dry_matter_abs_difference_pp",
            ["dry_matter_pct", "dry_matter_pct_from_weights"],
            "abs(dry_matter_pct - dry_matter_pct_from_weights)",
            "=ABS([@dry_matter_pct]-[@dry_matter_pct_from_weights])",
            "percentage points",
        ),
        formula(
            "dry_matter_relative_difference",
            "dry_matter_calculated.csv",
            "dry_matter_relative_difference",
            ["dry_matter_abs_difference_pp", "dry_matter_pct"],
            "dry_matter_abs_difference_pp / abs(dry_matter_pct)",
            "=[@dry_matter_abs_difference_pp]/ABS([@dry_matter_pct])",
            "ratio",
        ),
        formula(
            "dry_matter_issue",
            "dry_matter_calculated.csv",
            "dry_matter_issue",
            [
                "dry_matter_abs_difference_pp",
                "dry_matter_relative_difference",
            ],
            "dry_matter_abs_difference_pp >= 5 and dry_matter_relative_difference >= 0.20",
            "=AND([@dry_matter_abs_difference_pp]>=5,[@dry_matter_relative_difference]>=0.20)",
            None,
        ),
        formula(
            "biomass_recorded_dm",
            "dry_matter_calculated.csv",
            "biomass_recorded_dm_kg_ha",
            ["green_weight_1m_g", "dry_matter_pct", "biomass_sample_area_m2"],
            "green_weight_1m_g * (dry_matter_pct / 100) * 10 / biomass_sample_area_m2",
            "=[@green_weight_1m_g]*([@dry_matter_pct]/100)*10/biomass_sample_area_m2",
            "kg DM/ha",
        ),
        formula(
            "biomass_ratio_dm",
            "dry_matter_calculated.csv",
            "biomass_ratio_dm_kg_ha",
            [
                "green_weight_1m_g",
                "dry_matter_pct_from_weights",
                "biomass_sample_area_m2",
            ],
            "green_weight_1m_g * (dry_matter_pct_from_weights / 100) * 10 / biomass_sample_area_m2",
            "=[@green_weight_1m_g]*([@dry_matter_pct_from_weights]/100)*10/biomass_sample_area_m2",
            "kg DM/ha",
        ),
        formula(
            "biomass_recalculation_difference",
            "dry_matter_calculated.csv",
            "biomass_recalculation_difference_kg_ha",
            ["biomass_recorded_dm_kg_ha", "biomass_reported_kg_ha"],
            "biomass_recorded_dm_kg_ha - biomass_reported_kg_ha",
            "=[@biomass_recorded_dm_kg_ha]-[@biomass_reported_kg_ha]",
            "kg DM/ha",
        ),
        formula(
            "tiller_density",
            "dry_matter_calculated.csv",
            "tiller_density_m2",
            ["tillers_30_cm", "row_spacing_cm"],
            "tillers_30_cm * (100 / 30) * (100 / row_spacing_cm)",
            "=[@tillers_30_cm]*(100/30)*(100/row_spacing_cm)",
            "tillers/m2",
        ),
        formula(
            "thousand_seed_weight",
            "harvest_calculated.csv",
            "w1000_g",
            ["w100_rep1_g", "w100_rep2_g", "w100_rep3_g"],
            "mean(w100_rep1_g, w100_rep2_g, w100_rep3_g) * 10",
            "=AVERAGE([@w100_rep1_g]:[@w100_rep3_g])*10",
            "g",
        ),
        formula(
            "panicle_density",
            "harvest_calculated.csv",
            "panicle_density_m2",
            ["panicle_count", "harvest_sample_area_m2"],
            "panicle_count / harvest_sample_area_m2",
            "=[@panicle_count]/harvest_sample_area_m2",
            "panicles/m2",
        ),
        formula(
            "dirty_yield",
            "harvest_calculated.csv",
            "dirty_yield_kg_ha",
            ["dirty_seed_mass_g", "harvest_sample_area_m2"],
            "dirty_seed_mass_g * 10 / harvest_sample_area_m2",
            "=[@dirty_seed_mass_g]*10/harvest_sample_area_m2",
            "kg/ha",
        ),
        formula(
            "clean_yield",
            "harvest_calculated.csv",
            "clean_yield_kg_ha",
            ["clean_seed_mass_g", "harvest_sample_area_m2"],
            "clean_seed_mass_g * 10 / harvest_sample_area_m2",
            "=[@clean_seed_mass_g]*10/harvest_sample_area_m2",
            "kg/ha",
        ),
        formula(
            "clean_recovery",
            "harvest_calculated.csv",
            "clean_recovery",
            ["clean_seed_mass_g", "dirty_seed_mass_g"],
            "clean_seed_mass_g / dirty_seed_mass_g",
            "=[@clean_seed_mass_g]/[@dirty_seed_mass_g]",
            "ratio",
        ),
        formula(
            "cleaning_loss",
            "harvest_calculated.csv",
            "cleaning_loss_pct",
            ["clean_recovery"],
            "100 * (1 - clean_recovery)",
            "=100*(1-[@clean_recovery])",
            "%",
        ),
        formula(
            "estimated_seed_count",
            "harvest_calculated.csv",
            "estimated_seed_count",
            ["clean_seed_mass_g", "w1000_g"],
            "1000 * clean_seed_mass_g / w1000_g",
            "=1000*[@clean_seed_mass_g]/[@w1000_g]",
            "seeds",
        ),
        formula(
            "estimated_seeds_per_panicle",
            "harvest_calculated.csv",
            "estimated_seeds_per_panicle",
            ["estimated_seed_count", "panicle_count"],
            "estimated_seed_count / panicle_count",
            "=[@estimated_seed_count]/[@panicle_count]",
            "seeds/panicle",
        ),
        formula(
            "harvest_index",
            "harvest_calculated.csv",
            "harvest_index_pct",
            ["clean_yield_kg_ha", "final_biomass_kg_ha"],
            "100 * clean_yield_kg_ha / final_biomass_kg_ha",
            "=100*[@clean_yield_kg_ha]/[@final_biomass_kg_ha]",
            "%",
        ),
        *[
            formula(
                f"rcbd_missing_{prefix}",
                "missing_quality_estimate.csv",
                f"{prefix}_pct_estimated",
                [
                    f"{prefix}_block_total",
                    f"{prefix}_treatment_total",
                    f"{prefix}_grand_total",
                    "treatment_count",
                    "block_count",
                ],
                (
                    f"(treatment_count * {prefix}_block_total + "
                    f"block_count * {prefix}_treatment_total - "
                    f"{prefix}_grand_total) / "
                    "((treatment_count - 1) * (block_count - 1))"
                ),
                (
                    f"=(treatment_count*[@{prefix}_block_total]+"
                    f"block_count*[@{prefix}_treatment_total]-"
                    f"[@{prefix}_grand_total])/"
                    "((treatment_count-1)*(block_count-1))"
                ),
                "%",
            )
            for prefix in ("n", "adf", "ndf")
        ],
        *[
            formula(
                f"quality_{prefix}_estimate_lookup",
                "quality_calculated.csv",
                f"estimated_{prefix}_pct",
                ["sample_id", "measurement_status", f"{prefix}_pct_estimated"],
                (
                    f"if measurement_status == 'estimated', look up "
                    f"{prefix}_pct_estimated by sample_id; otherwise null"
                ),
                (
                    f'=IF([@measurement_status]="estimated",'
                    f"XLOOKUP([@sample_id],MissingQuality[sample_id],"
                    f'MissingQuality[{prefix}_pct_estimated]),"")'
                ),
                "%",
            )
            for prefix in ("n", "adf", "ndf")
        ],
        formula(
            "n_accumulated",
            "quality_calculated.csv",
            "n_accumulated_kg_ha",
            ["biomass_reported_kg_ha", "effective_n_pct"],
            "biomass_reported_kg_ha * effective_n_pct / 100",
            "=[@biomass_reported_kg_ha]*[@effective_n_pct]/100",
            "kg N/ha",
        ),
        formula(
            "critical_n",
            "quality_calculated.csv",
            "critical_n_pct",
            ["biomass_reported_kg_ha"],
            "4.8 * (biomass_reported_kg_ha / 1000) ** -0.32",
            "=4.8*([@biomass_reported_kg_ha]/1000)^(-0.32)",
            "%",
        ),
        formula(
            "nni",
            "quality_calculated.csv",
            "nni",
            ["effective_n_pct", "critical_n_pct"],
            "effective_n_pct / critical_n_pct",
            "=[@effective_n_pct]/[@critical_n_pct]",
            "ratio",
        ),
        formula(
            "missing_n_accumulated",
            "missing_quality_estimate.csv",
            "n_accumulated_kg_ha",
            ["biomass_kg_ha", "n_pct_estimated"],
            "biomass_kg_ha * n_pct_estimated / 100",
            "=[@biomass_kg_ha]*[@n_pct_estimated]/100",
            "kg N/ha",
        ),
        formula(
            "missing_critical_n",
            "missing_quality_estimate.csv",
            "critical_n_pct",
            ["biomass_kg_ha"],
            "4.8 * (biomass_kg_ha / 1000) ** -0.32",
            "=4.8*([@biomass_kg_ha]/1000)^(-0.32)",
            "%",
        ),
        formula(
            "missing_nni",
            "missing_quality_estimate.csv",
            "nni",
            ["n_pct_estimated", "critical_n_pct"],
            "n_pct_estimated / critical_n_pct",
            "=[@n_pct_estimated]/[@critical_n_pct]",
            "ratio",
        ),
    ]
    return {
        "schema_version": SCHEMA_VERSION,
        "evaluation_order": [
            "recorded_data",
            "dry_matter_calculations",
            "missing_quality_estimate",
            "quality_calculations",
            "harvest_calculations",
        ],
        "formulas": formulas,
        "note": (
            "Only formulas that materialize canonical calculated columns are retained; "
            "the historical workbook remains the archive of cell-level formulas."
        ),
    }


def _dataset_specs() -> dict[str, dict[str, object]]:
    specs: dict[str, dict[str, object]] = {}

    def add(
        file: str,
        sheet_name_es: str,
        grain_es: str,
        primary_key: Sequence[str],
        columns: Mapping[str, dict[str, object]],
        *,
        foreign_keys: Sequence[Mapping[str, object]] = (),
        sort_by: Sequence[str] = (),
        role: str = "recorded",
    ) -> None:
        specs[file] = {
            "future_sheet_name_es": sheet_name_es,
            "grain_es": grain_es,
            "role": role,
            "primary_key": list(primary_key),
            "foreign_keys": list(foreign_keys),
            "sort_by": list(sort_by or primary_key),
            "columns": dict(columns),
        }

    add(
        "experimental_design.csv",
        "Diseño experimental",
        "Una fila por posición de parcela dentro de cada bloque.",
        ["block", "plot_position"],
        {
            "block": _column("Bloque", "string"),
            "plot_position": _column("Posición de parcela", "integer"),
            "treatment": _column("Tratamiento", "string"),
        },
        sort_by=["block", "plot_position"],
    )
    add(
        "experiment_metadata.csv",
        "Información del ensayo",
        "Una fila por parámetro del ensayo.",
        ["parameter"],
        {
            "parameter": _column("Parámetro", "string"),
            "value": _column("Valor", "string"),
            "value_type": _column("Tipo de valor", "string"),
            "unit": _column("Unidad", "string", nullable=True),
            "legacy_source": _column("Referencia histórica", "string"),
        },
    )
    add(
        "field_timeline.csv",
        "Cronograma experimental",
        "Una fila por evento de campo atómico.",
        ["event_id"],
        {
            "event_id": _column("Identificador del evento", "string"),
            "date": _column("Fecha", "date"),
            "event_type": _column("Tipo de evento", "string"),
            "scope": _column("Alcance", "string"),
            "treatment": _column("Tratamiento", "string", nullable=True),
            "application_number": _column(
                "Número de aplicación", "integer", nullable=True
            ),
            "growth_stage": _column("Estado fenológico", "string", nullable=True),
            "description": _column("Descripción", "string"),
            "within_day_order_known": _column(
                "Orden intradía conocido", "boolean", nullable=True
            ),
        },
        sort_by=["date", "event_id"],
    )
    add(
        "soil_analysis.csv",
        "Análisis de suelo",
        "Una fila por determinación de suelo.",
        ["parameter"],
        {
            "parameter": _column("Parámetro", "string"),
            "value": _column("Valor", "number"),
            "unit": _column("Unidad", "string", nullable=True),
        },
    )
    add(
        "field_management.csv",
        "Manejo del cultivo",
        "Una fila por mes calendario.",
        ["year", "month"],
        {
            "year": _column("Año", "integer"),
            "month": _column("Mes", "integer"),
            "month_label": _column("Nombre del mes", "string"),
            "activity": _column("Actividad de manejo", "string", nullable=True),
        },
        sort_by=["year", "month"],
    )
    add(
        "water_inputs.csv",
        "Aportes de agua",
        "Una fila por mes calendario.",
        ["year", "month"],
        {
            "year": _column("Año", "integer"),
            "month": _column("Mes", "integer"),
            "month_label": _column("Nombre del mes", "string"),
            "irrigation_mm": _column("Riego suplementario", "number", unit="mm"),
            "rainfall_mm": _column("Precipitación", "number", unit="mm"),
        },
        sort_by=["year", "month"],
    )
    add(
        "dry_matter_recorded.csv",
        "Materia seca - datos registrados",
        "Una fila por observación de biomasa o macollos.",
        ["observation_id"],
        {
            "observation_id": _column("Identificador de observación", "string"),
            "date": _column("Fecha", "date"),
            "sample_id": _column("Muestra", "integer", nullable=True),
            "sector": _column("Sector hídrico", "string"),
            "treatment": _column("Tratamiento", "string"),
            "block": _column("Bloque", "string"),
            "tillers_30_cm": _column(
                "Macollos en 30 cm", "number", unit="macollos/30 cm", nullable=True
            ),
            "green_weight_1m_g": _column(
                "Peso verde en un metro", "number", unit="g", nullable=True
            ),
            "green_weight_sample_g": _column(
                "Peso verde de la muestra", "number", unit="g", nullable=True
            ),
            "tray_weight_g": _column(
                "Peso de bandeja", "number", unit="g", nullable=True
            ),
            "dry_weight_g": _column("Peso seco", "number", unit="g", nullable=True),
            "dry_matter_pct": _column(
                "Materia seca registrada", "number", unit="%", nullable=True
            ),
        },
        sort_by=["date", "sector", "treatment", "block"],
    )
    add(
        "dry_matter_calculated.csv",
        "Materia seca - cálculos",
        "Una fila calculada por observación de materia seca.",
        ["observation_id"],
        {
            "observation_id": _column("Identificador de observación", "string"),
            "dry_matter_pct_from_weights": _column(
                "Materia seca reconstruida", "number", unit="%", nullable=True
            ),
            "dry_matter_abs_difference_pp": _column(
                "Diferencia absoluta de materia seca",
                "number",
                unit="puntos porcentuales",
                nullable=True,
            ),
            "dry_matter_relative_difference": _column(
                "Diferencia relativa de materia seca",
                "number",
                unit="proporción",
                nullable=True,
            ),
            "dry_matter_issue": _column(
                "Registro de materia seca a verificar", "boolean"
            ),
            "biomass_reported_kg_ha": _column(
                "Biomasa informada en el libro histórico",
                "number",
                unit="kg MS/ha",
                nullable=True,
            ),
            "biomass_recorded_dm_kg_ha": _column(
                "Biomasa recalculada con materia seca registrada",
                "number",
                unit="kg MS/ha",
                nullable=True,
            ),
            "biomass_ratio_dm_kg_ha": _column(
                "Biomasa recalculada con materia seca reconstruida",
                "number",
                unit="kg MS/ha",
                nullable=True,
            ),
            "biomass_recalculation_difference_kg_ha": _column(
                "Diferencia entre biomasa recalculada e informada",
                "number",
                unit="kg MS/ha",
                nullable=True,
            ),
            "tiller_density_m2": _column(
                "Densidad de macollos", "number", unit="macollos/m2", nullable=True
            ),
        },
        foreign_keys=[
            {"columns": ["observation_id"], "references": "dry_matter_recorded.csv"}
        ],
        role="calculated",
    )
    add(
        "harvest_recorded.csv",
        "Rendimiento - datos registrados",
        "Una fila por parcela cosechada.",
        ["observation_id"],
        {
            "observation_id": _column("Identificador de observación", "string"),
            "date": _column("Fecha", "date"),
            "sample_id": _column("Muestra", "integer"),
            "sector": _column("Sector hídrico", "string"),
            "treatment": _column("Tratamiento", "string"),
            "block": _column("Bloque", "string"),
            "panicle_count": _column(
                "Panojas en el área cosechada", "number", unit="panojas"
            ),
            "dirty_seed_mass_g": _column(
                "Peso de semilla sin limpiar", "number", unit="g"
            ),
            "clean_seed_mass_g": _column("Peso de semilla limpia", "number", unit="g"),
            "w100_rep1_g": _column(
                "Peso de 100 semillas - réplica 1", "number", unit="g"
            ),
            "w100_rep2_g": _column(
                "Peso de 100 semillas - réplica 2", "number", unit="g"
            ),
            "w100_rep3_g": _column(
                "Peso de 100 semillas - réplica 3", "number", unit="g"
            ),
        },
        sort_by=["sector", "block", "treatment"],
    )
    add(
        "harvest_calculated.csv",
        "Rendimiento - cálculos",
        "Una fila calculada por parcela cosechada.",
        ["observation_id"],
        {
            "observation_id": _column("Identificador de observación", "string"),
            "w1000_g": _column("Peso de mil semillas", "number", unit="g"),
            "panicle_density_m2": _column(
                "Densidad de panojas", "number", unit="panojas/m2"
            ),
            "dirty_yield_kg_ha": _column(
                "Rendimiento sin limpiar", "number", unit="kg/ha"
            ),
            "clean_yield_kg_ha": _column("Rendimiento limpio", "number", unit="kg/ha"),
            "clean_recovery": _column(
                "Recuperación de limpieza", "number", unit="proporción"
            ),
            "cleaning_loss_pct": _column("Merma de limpieza", "number", unit="%"),
            "estimated_seed_count": _column(
                "Número estimado de semillas", "number", unit="semillas"
            ),
            "estimated_seeds_per_panicle": _column(
                "Semillas estimadas por panoja", "number", unit="semillas/panoja"
            ),
            "harvest_index_pct": _column("Índice de cosecha", "number", unit="%"),
        },
        foreign_keys=[
            {"columns": ["observation_id"], "references": "harvest_recorded.csv"}
        ],
        role="calculated",
    )
    add(
        "quality_recorded.csv",
        "Calidad - datos registrados",
        "Una fila por muestra de laboratorio o muestra faltante identificada.",
        ["sample_id"],
        {
            "sample_id": _column("Muestra", "integer"),
            "date": _column("Fecha", "date"),
            "sector": _column("Sector hídrico", "string"),
            "treatment": _column("Tratamiento", "string"),
            "block": _column("Bloque", "string"),
            "lab_registration": _column(
                "Registro de laboratorio", "integer", nullable=True
            ),
            "n_pct": _column("Nitrógeno medido", "number", unit="%", nullable=True),
            "adf_pct": _column(
                "Fibra detergente ácido medida", "number", unit="%", nullable=True
            ),
            "ndf_pct": _column(
                "Fibra detergente neutro medida", "number", unit="%", nullable=True
            ),
            "measurement_status": _column("Estado de la medición", "string"),
            "data_origin": _column("Origen del dato", "string"),
        },
        foreign_keys=[
            {"columns": ["sample_id"], "references": "dry_matter_recorded.csv"}
        ],
        sort_by=["sample_id"],
    )
    add(
        "quality_calculated.csv",
        "Calidad - cálculos",
        "Una fila calculada por muestra de calidad.",
        ["sample_id"],
        {
            "sample_id": _column("Muestra", "integer"),
            "estimated_n_pct": _column(
                "Nitrógeno estimado", "number", unit="%", nullable=True
            ),
            "estimated_adf_pct": _column(
                "Fibra detergente ácido estimada", "number", unit="%", nullable=True
            ),
            "estimated_ndf_pct": _column(
                "Fibra detergente neutro estimada", "number", unit="%", nullable=True
            ),
            "n_accumulated_kg_ha": _column(
                "Nitrógeno acumulado", "number", unit="kg N/ha"
            ),
            "critical_n_pct": _column(
                "Concentración crítica de nitrógeno", "number", unit="%"
            ),
            "nni": _column(
                "Índice de nutrición nitrogenada", "number", unit="proporción"
            ),
            "calculation_status": _column("Procedencia del cálculo", "string"),
        },
        foreign_keys=[{"columns": ["sample_id"], "references": "quality_recorded.csv"}],
        role="calculated",
        sort_by=["sample_id"],
    )
    add(
        "missing_quality_estimate.csv",
        "Estimación de calidad faltante",
        "Una fila para la muestra de calidad estimada por DBCA.",
        ["sample_id"],
        {
            "sample_id": _column("Muestra", "integer"),
            "date": _column("Fecha", "date"),
            "sector": _column("Sector hídrico", "string"),
            "treatment": _column("Tratamiento", "string"),
            "block": _column("Bloque", "string"),
            "biomass_kg_ha": _column("Biomasa", "number", unit="kg MS/ha"),
            "treatment_count": _column("Número de tratamientos", "integer"),
            "block_count": _column("Número de bloques", "integer"),
            "n_block_total": _column("Total de N del bloque", "number", unit="%"),
            "n_treatment_total": _column(
                "Total de N del tratamiento", "number", unit="%"
            ),
            "n_grand_total": _column("Total general de N", "number", unit="%"),
            "n_pct_estimated": _column("Nitrógeno estimado", "number", unit="%"),
            "adf_block_total": _column("Total de FDA del bloque", "number", unit="%"),
            "adf_treatment_total": _column(
                "Total de FDA del tratamiento", "number", unit="%"
            ),
            "adf_grand_total": _column("Total general de FDA", "number", unit="%"),
            "adf_pct_estimated": _column("FDA estimada", "number", unit="%"),
            "ndf_block_total": _column("Total de FDN del bloque", "number", unit="%"),
            "ndf_treatment_total": _column(
                "Total de FDN del tratamiento", "number", unit="%"
            ),
            "ndf_grand_total": _column("Total general de FDN", "number", unit="%"),
            "ndf_pct_estimated": _column("FDN estimada", "number", unit="%"),
            "n_accumulated_kg_ha": _column(
                "Nitrógeno acumulado estimado", "number", unit="kg N/ha"
            ),
            "critical_n_pct": _column(
                "Concentración crítica de nitrógeno", "number", unit="%"
            ),
            "nni": _column(
                "Índice de nutrición nitrogenada", "number", unit="proporción"
            ),
        },
        foreign_keys=[{"columns": ["sample_id"], "references": "quality_recorded.csv"}],
        role="calculated",
    )
    return specs


def _manifest_payload(source: Path) -> dict[str, object]:
    return {
        "schema_version": SCHEMA_VERSION,
        "authority": "canonical_csv_bundle",
        "migration": {
            "historical_source_file": (
                source.relative_to(PROJECT_ROOT).as_posix()
                if source.is_relative_to(PROJECT_ROOT)
                else str(source)
            ),
            "historical_source_sha256": sha256_file(source),
            "historical_workbook_role": "migration evidence only",
            "transformations": [
                "Normalized Diseño treatment typo MO to M0.",
                "Consolidated Ensayo schedule and activity notes into one atomic timeline.",
                "Fixed all experimental dates to 2025 and retained M3 application 2 on 2025-08-21.",
                "Separated recorded values from calculated materializations.",
                "Removed duplicated dry-matter and biomass fields from quality data.",
                "Dropped trivial water total rows; totals are validated from monthly inputs.",
                "Replaced the legacy circular quality-estimation graph with an acyclic DBCA calculation sequence.",
            ],
        },
        "datasets": _dataset_specs(),
    }


def _read_metadata_values(rows: Sequence[Mapping[str, object]]) -> dict[str, object]:
    values: dict[str, object] = {}
    for row in rows:
        parameter = cast(str, row["parameter"])
        value = row["value"]
        value_type = row["value_type"]
        if value_type == "integer":
            values[parameter] = int(cast(Any, value))
        elif value_type == "number":
            values[parameter] = float(cast(Any, value))
        else:
            values[parameter] = value
    return values


def _build_migration(source: Path, output_dir: Path) -> None:
    value_book = load_workbook(source, data_only=True, read_only=False)
    try:
        metadata_rows = _metadata_rows(value_book)
        metadata = _read_metadata_values(metadata_rows)
        dry_recorded, dry_calculated = _dry_matter_rows(value_book, metadata)
        biomass_by_sample = {
            cast(int, raw["sample_id"]): cast(float, calc["biomass_reported_kg_ha"])
            for raw, calc in zip(dry_recorded, dry_calculated, strict=True)
            if raw["sample_id"] is not None
            and calc["biomass_reported_kg_ha"] is not None
        }
        biomass_by_plot = {
            (
                cast(str, raw["sector"]),
                cast(str, raw["treatment"]),
                cast(str, raw["block"]),
            ): cast(float, calc["biomass_recorded_dm_kg_ha"])
            for raw, calc in zip(dry_recorded, dry_calculated, strict=True)
            if raw["date"] == "2025-11-12"
            and calc["biomass_recorded_dm_kg_ha"] is not None
        }
        harvest_recorded, harvest_calculated = _harvest_rows(
            value_book, metadata, biomass_by_plot
        )
        quality_recorded, quality_calculated, missing_estimate = _quality_rows(
            value_book, biomass_by_sample
        )
        datasets: dict[str, list[dict[str, object]]] = {
            "experimental_design.csv": _design_rows(value_book),
            "experiment_metadata.csv": metadata_rows,
            "field_timeline.csv": _timeline_rows(),
            "soil_analysis.csv": _soil_rows(value_book),
            "field_management.csv": _management_rows(value_book),
            "water_inputs.csv": _water_rows(value_book),
            "dry_matter_recorded.csv": dry_recorded,
            "dry_matter_calculated.csv": dry_calculated,
            "harvest_recorded.csv": harvest_recorded,
            "harvest_calculated.csv": harvest_calculated,
            "quality_recorded.csv": quality_recorded,
            "quality_calculated.csv": quality_calculated,
            "missing_quality_estimate.csv": missing_estimate,
        }
        expected = set(_dataset_specs())
        if set(datasets) != expected:
            raise AssertionError("Migration datasets and manifest schemas diverged.")

        output_dir.mkdir(parents=True, exist_ok=False)
        for filename, rows in datasets.items():
            _write_csv(output_dir / filename, rows)
        _write_json(output_dir / "formulas.json", _formula_payload())
        _write_json(output_dir / "manifest.json", _manifest_payload(source))
        (output_dir / "README.md").write_text(
            "# Datos canónicos del ensayo\n\n"
            "Los CSV de este directorio son la única fuente editable y analítica. "
            "Los archivos `*_recorded.csv` contienen mediciones o metadatos; los "
            "archivos `*_calculated.csv` son materializaciones comprobables de las "
            "reglas declaradas en `formulas.json`.\n\n"
            "`manifest.json` define tipos, unidades, claves, relaciones y rótulos "
            "descriptivos en español para una futura reconstrucción del libro. El XLSX "
            "histórico se conserva únicamente como evidencia de la migración.\n\n"
            "Para validar el conjunto sin ejecutar los análisis completos:\n\n"
            "```bash\nuv run festuca-validate-data\n```\n\n"
            "No vuelva a ejecutar la migración sobre datos revisados manualmente salvo "
            "que realmente quiera reemplazarlos y use `--force`.\n",
            encoding="utf-8",
        )
    finally:
        value_book.close()


def _directory_files(path: Path) -> dict[str, bytes]:
    return {
        child.relative_to(path).as_posix(): child.read_bytes()
        for child in sorted(path.rglob("*"))
        if child.is_file()
    }


def migrate_workbook(
    source: Path, output_dir: Path, *, check: bool = False, force: bool = False
) -> None:
    source = source.resolve()
    output_dir = output_dir.resolve()
    if not source.is_file():
        raise FileNotFoundError(f"Workbook does not exist: {source}")
    if source.suffix.casefold() != ".xlsx":
        raise ValueError(f"Expected an XLSX migration source: {source}")
    lock_file = source.with_name(f"~${source.name}")
    if lock_file.exists():
        raise RuntimeError(
            f"Close Microsoft Excel before migrating the workbook: {lock_file}"
        )
    if check and force:
        raise ValueError("--check and --force cannot be used together.")

    output_dir.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(
        prefix=f".{output_dir.name}-migration-", dir=output_dir.parent
    ) as temporary_parent:
        candidate = Path(temporary_parent) / output_dir.name
        _build_migration(source, candidate)
        if check:
            if not output_dir.is_dir():
                raise FileNotFoundError(
                    f"Canonical data directory does not exist: {output_dir}"
                )
            expected = _directory_files(candidate)
            actual = _directory_files(output_dir)
            if actual != expected:
                differing = sorted(set(actual) ^ set(expected))
                differing.extend(
                    name
                    for name in sorted(set(actual) & set(expected))
                    if actual[name] != expected[name]
                )
                raise RuntimeError(
                    "Canonical data differs from a fresh workbook migration: "
                    + ", ".join(dict.fromkeys(differing))
                )
            return

        if output_dir.exists() and not force:
            raise FileExistsError(
                f"Canonical data already exists: {output_dir}. Use --force to replace it."
            )
        if not output_dir.exists():
            os.replace(candidate, output_dir)
            return

        backup = Path(temporary_parent) / f"{output_dir.name}-previous"
        os.replace(output_dir, backup)
        try:
            os.replace(candidate, output_dir)
        except Exception:
            os.replace(backup, output_dir)
            raise
        shutil.rmtree(backup)


def main() -> None:
    arguments = parse_args()
    migrate_workbook(
        arguments.source,
        arguments.output_dir,
        check=arguments.check,
        force=arguments.force,
    )


if __name__ == "__main__":
    main()
