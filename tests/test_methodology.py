from __future__ import annotations

import ast
import json
import re
import unittest
from pathlib import Path
from types import SimpleNamespace
from typing import Any, TypedDict, cast
from unittest.mock import patch

import numpy as np
import pandas as pd
import statsmodels.formula.api as smf  # pyright: ignore[reportMissingTypeStubs]
from openpyxl import load_workbook

from festuca_analysis.longitudinal import (
    LongitudinalNotebook,
    _stable_seed,  # pyright: ignore[reportPrivateUsage]
)
from festuca_analysis.source_data import load_experiment_data, sha256_file
from festuca_analysis.statistics import (
    benjamini_hochberg,
    fit_mixedlm_best,
    likelihood_ratio,
    rcbd_missing_cell_estimate,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = PROJECT_ROOT / "sources" / "Datos_Ema_Serrana_INN.xlsx"
NOTEBOOKS = (
    PROJECT_ROOT / "festuca_estudio_longitudinal.ipynb",
    PROJECT_ROOT / "festuca_anexo_probabilistico.ipynb",
)


class NotebookCell(TypedDict, total=False):
    cell_type: str
    execution_count: int | None
    metadata: dict[str, object]
    outputs: list[object]
    source: str | list[str]


class NotebookDocument(TypedDict):
    cells: list[NotebookCell]
    metadata: dict[str, object]


def _read_notebook(path: Path) -> NotebookDocument:
    raw: object = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise TypeError(f"Notebook root must be an object: {path}")
    return cast(NotebookDocument, raw)


def _cell_source(cell: NotebookCell) -> str:
    source = cell.get("source", "")
    return "".join(source) if isinstance(source, list) else source


def _festuca_metadata(notebook: NotebookDocument) -> dict[str, object]:
    metadata = notebook["metadata"].get("festuca")
    if not isinstance(metadata, dict):
        raise TypeError("Notebook is missing object metadata.festuca")
    return cast(dict[str, object], metadata)


class WorkbookSourceTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.data = load_experiment_data(WORKBOOK)

    def test_provenance_hash_is_computed_from_the_workbook(self) -> None:
        self.assertEqual(self.data.spec.source_sha256, sha256_file(WORKBOOK))

    def test_structured_schedule_is_read_from_ensayo(self) -> None:
        workbook = load_workbook(WORKBOOK, data_only=True, read_only=True)
        try:
            worksheet = workbook["Ensayo"]
            workbook_treatments = [
                str(worksheet.cell(row, 6).value).strip().upper() for row in range(2, 8)
            ]
        finally:
            workbook.close()
        self.assertEqual(
            self.data.spec.schedule["treatment"].tolist(),
            workbook_treatments,
        )
        self.assertEqual(
            self.data.spec.schedule["source_range"].tolist(),
            [f"F{row}:H{row}" for row in range(2, 8)],
        )

    def test_harvest_quantities_are_reconstructed_from_primitives(self) -> None:
        harvest = self.data.harvest
        area = self.data.spec.harvest_sample_area_m2
        np.testing.assert_allclose(
            harvest["clean_yield_kg_ha"],
            harvest["clean_mass_g"] * 10.0 / area,
        )
        np.testing.assert_allclose(
            harvest["panicle_density_m2"],
            harvest["panicle_count"] / area,
        )
        np.testing.assert_allclose(
            harvest["w1000_g"],
            harvest[["w100_1_g", "w100_2_g", "w100_3_g"]].mean(axis=1) * 10.0,
        )
        np.testing.assert_allclose(
            harvest["estimated_seeds_per_panicle"],
            1000.0
            * harvest["clean_mass_g"]
            / harvest["w1000_g"]
            / harvest["panicle_count"],
        )

    def test_quality_estimates_are_identified_and_excluded_from_primary_n(self) -> None:
        frame = self.data.longitudinal
        estimated = frame["quality_status"].eq("estimated_in_workbook")
        measured = frame["quality_status"].eq("recorded")
        self.assertGreater(int(estimated.sum()), 0)
        self.assertTrue(frame.loc[estimated, "n_pct"].isna().all())
        np.testing.assert_allclose(
            frame.loc[measured, "n_pct"],
            frame.loc[measured, "n_pct_recorded"],
        )
        self.assertTrue(
            frame.loc[estimated, "n_pct_cell_status"].eq("estimated_in_workbook").all()
        )

    def test_dry_matter_flags_follow_the_declared_dynamic_rule(self) -> None:
        frame = self.data.longitudinal
        expected = (
            frame["dm_abs_difference_pp"].ge(5.0)
            & frame["dm_relative_difference"].ge(0.20)
        ).fillna(False)
        pd.testing.assert_series_equal(
            frame["dm_issue"].reset_index(drop=True),
            expected.reset_index(drop=True),
            check_names=False,
        )

    def test_lineage_distinguishes_recorded_calculated_and_estimated_values(
        self,
    ) -> None:
        statuses = " | ".join(self.data.variable_lineage["status"].astype(str))
        for expected in (
            "recorded",
            "calculated_in_workbook",
            "estimated_in_workbook",
            "analysis_derived",
        ):
            self.assertIn(expected, statuses)


class StatisticalUtilityTests(unittest.TestCase):
    def test_benjamini_hochberg_preserves_order_and_monotonicity(self) -> None:
        adjusted = benjamini_hochberg([0.04, 0.001, 0.03, 0.20])
        np.testing.assert_allclose(
            adjusted,
            [0.05333333333333334, 0.004, 0.05333333333333334, 0.20],
        )

    def test_rcbd_missing_cell_estimate_uses_correct_denominators(self) -> None:
        frame = pd.DataFrame(
            {
                "treatment": ["A"] * 3 + ["B"] * 3 + ["C"] * 3 + ["D"] * 3,
                "block": ["R1", "R2", "R3"] * 4,
                "value": [
                    10.0,
                    12.0,
                    np.nan,
                    13.0,
                    12.0,
                    15.0,
                    9.0,
                    11.0,
                    10.0,
                    16.0,
                    15.0,
                    17.0,
                ],
            }
        )
        estimate = rcbd_missing_cell_estimate(
            frame,
            value_column="value",
            missing_treatment="A",
            missing_block="R3",
            r_blocks=3,
            t_treatments=4,
        )
        self.assertAlmostEqual(estimate, 12.333333333333334)

    def test_mixed_model_selection_uses_best_converged_fit(self) -> None:
        fits = {
            "first": SimpleNamespace(llf=-12.0, converged=False),
            "second": SimpleNamespace(llf=-10.0, converged=True),
            "third": SimpleNamespace(llf=-9.0, converged=True),
        }

        def fake_fit(**kwargs: Any) -> SimpleNamespace:
            return fits[str(kwargs["method"])]

        fake_model = SimpleNamespace(fit=fake_fit)
        frame = pd.DataFrame({"y": [0.0, 1.0], "plot_id": ["a", "b"]})
        with patch(
            "festuca_analysis.statistics.smf.mixedlm",
            return_value=fake_model,
        ):
            selected = fit_mixedlm_best(
                "y ~ 1",
                frame,
                methods=("first", "second", "third"),
            )
        self.assertIs(selected, fits["third"])
        self.assertEqual(selected._audit_optimizer, "third")
        self.assertEqual(selected._audit_selection, "best_converged")

    def test_mixed_model_selection_fails_closed_without_convergence(self) -> None:
        fits = {
            "first": SimpleNamespace(llf=-12.0, converged=False),
            "second": SimpleNamespace(llf=-10.0, converged=False),
        }

        def fake_fit(**kwargs: Any) -> SimpleNamespace:
            return fits[str(kwargs["method"])]

        fake_model = SimpleNamespace(fit=fake_fit)
        frame = pd.DataFrame({"y": [0.0, 1.0], "plot_id": ["a", "b"]})
        with (
            patch(
                "festuca_analysis.statistics.smf.mixedlm",
                return_value=fake_model,
            ),
            self.assertRaises(RuntimeError),
        ):
            fit_mixedlm_best(
                "y ~ 1",
                frame,
                methods=("first", "second"),
            )

    def test_likelihood_ratio_uses_parameter_difference(self) -> None:
        reduced = SimpleNamespace(llf=-10.0, df_modelwc=4)
        full = SimpleNamespace(llf=-7.0, df_modelwc=6)
        result = likelihood_ratio(reduced, full)
        self.assertEqual(result.statistic, 6.0)
        self.assertEqual(result.degrees_freedom, 2)
        self.assertGreater(result.p_asymptotic, 0.0)
        self.assertLess(result.p_asymptotic, 0.1)

    def test_partial_correlation_p_value_comes_from_the_full_regression(self) -> None:
        rng = np.random.default_rng(7)
        size = 80
        frame = pd.DataFrame(
            {
                "treatment": np.resize(["A", "B", "C", "D"], size),
                "block": np.resize(["R1", "R2", "R3", "R4"], size),
                "x": rng.normal(size=size),
            }
        )
        treatment_effect = frame["treatment"].map(
            {"A": -0.5, "B": 0.0, "C": 0.25, "D": 0.75}
        )
        frame["y"] = (
            0.65 * frame["x"]
            + treatment_effect
            + rng.normal(
                scale=0.7,
                size=size,
            )
        )
        partial_r, p_value, n = (
            LongitudinalNotebook._partial_correlation_from_regression(  # pyright: ignore[reportPrivateUsage]
                frame,
                x="x",
                y="y",
                controls=["treatment", "block"],
            )
        )
        reference = (
            cast(Any, smf)
            .ols(
                "y ~ x + C(treatment) + C(block)",
                data=frame,
            )
            .fit()
        )
        t_value = float(reference.tvalues["x"])
        expected_r = np.sign(t_value) * np.sqrt(
            t_value**2 / (t_value**2 + reference.df_resid)
        )
        self.assertEqual(n, size)
        self.assertAlmostEqual(partial_r, expected_r)
        self.assertAlmostEqual(p_value, float(reference.pvalues["x"]))

    def test_stable_seed_is_process_independent(self) -> None:
        self.assertEqual(
            _stable_seed("Secano", "biomass", "raw", base=123),
            _stable_seed("Secano", "biomass", "raw", base=123),
        )
        self.assertNotEqual(
            _stable_seed("Secano", "biomass", "raw", base=123),
            _stable_seed("Riego", "biomass", "raw", base=123),
        )


class NotebookHygieneTests(unittest.TestCase):
    def test_notebooks_are_unexecuted_and_markdown_is_methods_only(self) -> None:
        for path in NOTEBOOKS:
            notebook = _read_notebook(path)
            festuca_metadata = _festuca_metadata(notebook)
            self.assertEqual(
                festuca_metadata["source_of_truth"],
                "sources/Datos_Ema_Serrana_INN.xlsx",
            )
            self.assertEqual(
                festuca_metadata["markdown_policy"],
                "mathematics_and_logic_only",
            )
            for cell in notebook["cells"]:
                if cell.get("cell_type") == "code":
                    self.assertIsNone(cell.get("execution_count"))
                    self.assertEqual(cell.get("outputs"), [])
                elif cell.get("cell_type") == "markdown":
                    tags = cell.get("metadata", {}).get("tags", [])
                    self.assertIsInstance(tags, list)
                    self.assertIn("methods-only", cast(list[object], tags))

    def test_notebook_code_cells_are_syntactically_valid(self) -> None:
        for path in NOTEBOOKS:
            notebook = _read_notebook(path)
            for index, cell in enumerate(notebook["cells"]):
                if cell.get("cell_type") != "code":
                    continue
                try:
                    ast.parse(_cell_source(cell), filename=f"{path.name}:cell-{index}")
                except SyntaxError as error:
                    self.fail(f"Invalid code in {path.name}, cell {index}: {error}")

    def test_markdown_contains_methods_not_embedded_results_tables(self) -> None:
        forbidden_patterns = {
            "ISO date": re.compile(r"\b20\d{2}-\d{2}-\d{2}\b"),
            "reported p-value": re.compile(r"\bp\s*=\s*0?\.\d+", re.IGNORECASE),
            "markdown data table": re.compile(r"\|\s*:?-{3,}:?\s*\|"),
        }
        for path in NOTEBOOKS:
            notebook = _read_notebook(path)
            markdown = "\n".join(
                _cell_source(cell)
                for cell in notebook["cells"]
                if cell.get("cell_type") == "markdown"
            )
            for label, pattern in forbidden_patterns.items():
                self.assertIsNone(
                    pattern.search(markdown),
                    msg=f"{path.name} contains an embedded {label}",
                )

    def test_notebooks_do_not_read_generated_csv_as_input(self) -> None:
        for path in NOTEBOOKS:
            notebook = _read_notebook(path)
            source = "\n".join(
                _cell_source(cell)
                for cell in notebook["cells"]
                if cell.get("cell_type") == "code"
            )
            self.assertNotIn("read_csv", source)
            self.assertNotIn("legacy_probabilistic_run", source)
            self.assertNotIn("model_b", source.casefold())

    def test_notebooks_use_the_installed_package_without_path_mutation(self) -> None:
        for path in NOTEBOOKS:
            notebook = _read_notebook(path)
            source = "\n".join(
                _cell_source(cell)
                for cell in notebook["cells"]
                if cell.get("cell_type") == "code"
            )
            self.assertNotIn("sys.path", source)
            self.assertNotIn("source_directory", source)

    def test_probabilistic_module_has_no_frozen_legacy_dependency(self) -> None:
        source = (PROJECT_ROOT / "src" / "festuca_analysis" / "annex.py").read_text(
            encoding="utf-8"
        )
        self.assertNotIn("legacy_probabilistic_run", source)
        self.assertNotIn("model_b", source.casefold())


if __name__ == "__main__":
    unittest.main()
